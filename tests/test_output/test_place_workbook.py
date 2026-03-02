"""Contract tests for PP-003 IMP-14 place workbook builder."""

from __future__ import annotations

import importlib
import sys
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import load_workbook

PROJECT_ROOT = Path(__file__).resolve().parents[2]
sys.path.insert(0, str(PROJECT_ROOT))

workbook_mod = importlib.import_module("scripts.exports.build_place_workbook")

KEY_YEARS = [2025, 2030, 2035, 2040, 2045, 2050, 2055]
COUNTIES = ["38001", "38003", "38005", "38007", "38009", "38011", "38013", "38015", "38017"]


def _write_county_reference_csv(root: Path) -> None:
    """Write minimal ND county reference file used by workbook county labels."""
    county_rows = []
    for county_fips in COUNTIES:
        county_rows.append(
            {
                "state_fips": "38",
                "county_fips": county_fips,
                "county_name": f"County {county_fips} County",
            }
        )
    county_csv = root / "data" / "raw" / "geographic" / "nd_counties.csv"
    county_csv.parent.mkdir(parents=True, exist_ok=True)
    pd.DataFrame(county_rows).to_csv(county_csv, index=False)


def _write_place_parquet(
    place_dir: Path,
    place_fips: str,
    scenario: str,
    tier: str,
    place_index: int,
) -> None:
    """Write synthetic place parquet matching tier schema."""
    rows: list[dict[str, Any]] = []

    if tier == "HIGH":
        age_groups = workbook_mod.HIGH_AGE_GROUPS
    elif tier == "MODERATE":
        age_groups = workbook_mod.MODERATE_AGE_GROUPS
    else:
        age_groups = []

    for year_index, year in enumerate(KEY_YEARS):
        if tier == "LOWER":
            rows.append(
                {
                    "year": year,
                    "population": float(500 + place_index * 10 + year_index * 5),
                }
            )
            continue

        for age_index, age_group in enumerate(age_groups):
            for sex_index, sex in enumerate(["Male", "Female"]):
                rows.append(
                    {
                        "year": year,
                        "age_group": age_group,
                        "sex": sex,
                        "population": float(100 + place_index * 7 + age_index * 3 + sex_index + year_index),
                    }
                )

    place_df = pd.DataFrame(rows)
    output_path = place_dir / f"nd_place_{place_fips}_projection_2025_2055_{scenario}.parquet"
    place_df.to_parquet(output_path, index=False)


def _write_synthetic_place_outputs(root: Path, scenario: str) -> None:
    """Create synthetic PP-003 scenario outputs (90 places + balance rows)."""
    place_dir = root / "data" / "projections" / scenario / "place"
    place_dir.mkdir(parents=True, exist_ok=True)

    summary_rows: list[dict[str, Any]] = []
    all_places: list[tuple[str, str, str, str]] = []

    for idx in range(1, 10):
        all_places.append((f"3810{idx:03d}", f"High Place {idx}", "HIGH", COUNTIES[(idx - 1) % len(COUNTIES)]))
    for idx in range(1, 10):
        all_places.append(
            (f"3820{idx:03d}", f"Moderate Place {idx}", "MODERATE", COUNTIES[(idx + 1) % len(COUNTIES)])
        )
    for idx in range(1, 73):
        all_places.append((f"3830{idx:03d}", f"Lower Place {idx}", "LOWER", COUNTIES[(idx + 2) % len(COUNTIES)]))

    for place_index, (place_fips, place_name, tier, county_fips) in enumerate(all_places, start=1):
        _write_place_parquet(
            place_dir=place_dir,
            place_fips=place_fips,
            scenario=scenario,
            tier=tier,
            place_index=place_index,
        )
        base_population = float(1000 + place_index * 25)
        final_population = float(base_population + 100 + place_index)
        growth_rate = (final_population - base_population) / base_population
        summary_rows.append(
            {
                "place_fips": place_fips,
                "name": place_name,
                "county_fips": county_fips,
                "level": "place",
                "row_type": "place",
                "confidence_tier": tier,
                "base_population": base_population,
                "final_population": final_population,
                "absolute_growth": final_population - base_population,
                "growth_rate": growth_rate,
                "base_share": 0.1,
                "final_share": 0.1,
                "processing_time": 0.01,
            }
        )

    # Include balance rows to verify workbook input filtering uses row_type == place.
    summary_rows.extend(
        [
            {
                "place_fips": "bal_38001",
                "name": "Balance of County 38001",
                "county_fips": "38001",
                "level": "place",
                "row_type": "balance_of_county",
                "confidence_tier": None,
                "base_population": 500.0,
                "final_population": 520.0,
                "absolute_growth": 20.0,
                "growth_rate": 0.04,
                "base_share": 0.2,
                "final_share": 0.2,
                "processing_time": 0.0,
            },
            {
                "place_fips": "bal_38003",
                "name": "Balance of County 38003",
                "county_fips": "38003",
                "level": "place",
                "row_type": "balance_of_county",
                "confidence_tier": None,
                "base_population": 700.0,
                "final_population": 730.0,
                "absolute_growth": 30.0,
                "growth_rate": 0.0428,
                "base_share": 0.3,
                "final_share": 0.3,
                "processing_time": 0.0,
            },
        ]
    )

    pd.DataFrame(summary_rows).to_csv(place_dir / "places_summary.csv", index=False)


def _find_header_row(ws: Any, header_value: str) -> int:
    """Return row index containing expected first-column header value."""
    for row_idx in range(1, 80):
        if ws.cell(row=row_idx, column=1).value == header_value:
            return row_idx
    raise AssertionError(f"Header '{header_value}' not found in worksheet {ws.title}")


def test_build_place_workbook_contract(monkeypatch, tmp_path: Path) -> None:
    """
    Build workbook from synthetic place outputs and verify IMP-14 contract.

    Contract checks include:
    - 21-sheet workbook structure
    - HIGH/MODERATE sheet table shapes
    - LOWER combined sheet row count + caveat
    - TOC hyperlinks
    - methodology text and output naming pattern
    """
    scenario = "baseline"
    _write_county_reference_csv(tmp_path)
    _write_synthetic_place_outputs(tmp_path, scenario=scenario)

    config = {
        "pipeline": {
            "projection": {"output_dir": "data/projections"},
            "export": {"output_dir": "data/exports"},
        },
        "place_projections": {"output": {"key_years": KEY_YEARS}},
    }

    monkeypatch.setattr(workbook_mod, "project_root", tmp_path)

    output_path = workbook_mod.build_workbook(
        scenario=scenario,
        config=config,
        date_stamp="20260301",
    )

    assert output_path.name == "nd_projections_baseline_places_20260301.xlsx"
    assert output_path.exists()

    wb = load_workbook(output_path, data_only=True)
    sheetnames = wb.sheetnames

    assert len(sheetnames) == 21
    assert sheetnames[0] == "Table of Contents"
    assert "LOWER Tier" in sheetnames
    assert "Methodology" in sheetnames
    assert sum(name.startswith("HIGH - ") for name in sheetnames) == 9
    assert sum(name.startswith("MODERATE - ") for name in sheetnames) == 9

    # TOC must link all 90 places.
    toc_ws = wb["Table of Contents"]
    toc_header_row = _find_header_row(toc_ws, "Place Name")
    toc_rows = []
    row = toc_header_row + 1
    while toc_ws.cell(row=row, column=1).value:
        toc_rows.append(row)
        row += 1
    assert len(toc_rows) == 90
    for row_idx in toc_rows:
        link = toc_ws.cell(row=row_idx, column=1).hyperlink
        assert link is not None
        assert str(link.target).startswith("#'")
        assert str(link.target).endswith("'!A1")

    # HIGH tier: 18 age rows and 2 sex columns per key year.
    first_high_name = next(name for name in sheetnames if name.startswith("HIGH - "))
    high_ws = wb[first_high_name]
    high_header_row = _find_header_row(high_ws, "Age Group")
    expected_high_headers = ["Age Group"] + [value for year in KEY_YEARS for value in (f"{year} Male", f"{year} Female")]
    actual_high_headers = [
        high_ws.cell(row=high_header_row, column=col_idx).value for col_idx in range(1, len(expected_high_headers) + 1)
    ]
    assert actual_high_headers == expected_high_headers
    high_age_labels = [high_ws.cell(row=high_header_row + 1 + idx, column=1).value for idx in range(18)]
    assert high_age_labels == workbook_mod.HIGH_AGE_GROUPS

    # MODERATE tier: 6 broad age rows and 2 sex columns per key year.
    first_mod_name = next(name for name in sheetnames if name.startswith("MODERATE - "))
    mod_ws = wb[first_mod_name]
    mod_header_row = _find_header_row(mod_ws, "Age Group")
    expected_mod_headers = ["Age Group"] + [value for year in KEY_YEARS for value in (f"{year} Male", f"{year} Female")]
    actual_mod_headers = [
        mod_ws.cell(row=mod_header_row, column=col_idx).value for col_idx in range(1, len(expected_mod_headers) + 1)
    ]
    assert actual_mod_headers == expected_mod_headers
    mod_age_labels = [mod_ws.cell(row=mod_header_row + 1 + idx, column=1).value for idx in range(6)]
    assert mod_age_labels == workbook_mod.MODERATE_AGE_GROUPS

    # LOWER tier sheet: 72 rows and caveat header.
    lower_ws = wb["LOWER Tier"]
    lower_header_row = _find_header_row(lower_ws, "Place Name")
    lower_rows = []
    row = lower_header_row + 1
    while lower_ws.cell(row=row, column=1).value:
        lower_rows.append(row)
        row += 1
    assert len(lower_rows) == 72
    assert [lower_ws.cell(row=lower_header_row, column=4 + idx).value for idx in range(len(KEY_YEARS))] == [
        str(year) for year in KEY_YEARS
    ]
    caveat_cell = None
    for row_idx in range(1, lower_header_row):
        value = lower_ws.cell(row=row_idx, column=1).value
        if isinstance(value, str) and "wider uncertainty bands" in value:
            caveat_cell = lower_ws.cell(row=row_idx, column=1)
            break
    assert caveat_cell is not None
    assert caveat_cell.font.bold is True

    methodology_ws = wb["Methodology"]
    method_text = " ".join(
        str(methodology_ws.cell(row=row_idx, column=1).value or "") for row_idx in range(1, 60)
    )
    assert "Share-of-county trending method" in method_text
    assert "ADR-033 accepted, implemented 2026-03-01" in method_text
    assert "Winning backtest variant: B-II" in method_text
    assert "IMP-19 end-to-end validation passed" in method_text

    # HU Comparison sheet must NOT be present (no HU parquet written).
    assert "HU Comparison" not in sheetnames


def _write_synthetic_hu_projections(root: Path, scenario: str) -> None:
    """Write synthetic housing-unit projections for HIGH and MODERATE places."""
    place_dir = root / "data" / "projections" / scenario / "place"
    place_dir.mkdir(parents=True, exist_ok=True)

    hu_years = [2025, 2030, 2035]
    hu_rows: list[dict[str, Any]] = []

    # Write HU projections for all 9 HIGH and 9 MODERATE places.
    for idx in range(1, 10):
        place_fips = f"3810{idx:03d}"
        for year in hu_years:
            hu_rows.append({
                "place_fips": place_fips,
                "year": year,
                "hu_projected": float(400 + idx * 20 + (year - 2025)),
                "pph_projected": 2.5,
                "population_hu": float((400 + idx * 20 + (year - 2025)) * 2.5),
                "method": "hu_log_linear",
            })
    for idx in range(1, 10):
        place_fips = f"3820{idx:03d}"
        for year in hu_years:
            hu_rows.append({
                "place_fips": place_fips,
                "year": year,
                "hu_projected": float(200 + idx * 15 + (year - 2025)),
                "pph_projected": 2.3,
                "population_hu": float((200 + idx * 15 + (year - 2025)) * 2.3),
                "method": "hu_log_linear",
            })

    pd.DataFrame(hu_rows).to_parquet(
        place_dir / "housing_unit_projections.parquet", index=False,
    )


def test_hu_comparison_sheet_present(monkeypatch, tmp_path: Path) -> None:
    """
    When HU projections are available, the HU Comparison sheet is added.

    Verifies:
    - 22-sheet workbook (21 base + HU Comparison)
    - HU Comparison sheet is between LOWER Tier and Methodology
    - Correct header structure: Place Name, Tier, Place FIPS, then year groups
    - 18 data rows (9 HIGH + 9 MODERATE)
    - Explanatory note is present
    - Values are numeric and Diff/% Diff are computed correctly
    """
    scenario = "baseline"
    _write_county_reference_csv(tmp_path)
    _write_synthetic_place_outputs(tmp_path, scenario=scenario)
    _write_synthetic_hu_projections(tmp_path, scenario=scenario)

    config = {
        "pipeline": {
            "projection": {"output_dir": "data/projections"},
            "export": {"output_dir": "data/exports"},
        },
        "place_projections": {"output": {"key_years": KEY_YEARS}},
    }

    monkeypatch.setattr(workbook_mod, "project_root", tmp_path)

    output_path = workbook_mod.build_workbook(
        scenario=scenario,
        config=config,
        date_stamp="20260301",
    )

    assert output_path.exists()
    wb = load_workbook(output_path, data_only=True)
    sheetnames = wb.sheetnames

    # HU Comparison should be present, making 22 sheets total.
    assert len(sheetnames) == 22
    assert "HU Comparison" in sheetnames

    # HU Comparison must come between LOWER Tier and Methodology.
    hu_idx = sheetnames.index("HU Comparison")
    lower_idx = sheetnames.index("LOWER Tier")
    method_idx = sheetnames.index("Methodology")
    assert lower_idx < hu_idx < method_idx

    hu_ws = wb["HU Comparison"]

    # Title and note checks.
    assert hu_ws.cell(row=1, column=1).value == "Housing-Unit Method Comparison"

    note_found = False
    for row_idx in range(1, 15):
        cell_value = hu_ws.cell(row=row_idx, column=1).value
        if isinstance(cell_value, str) and "complementary cross-check" in cell_value:
            note_found = True
            break
    assert note_found, "Explanatory note not found in HU Comparison sheet"

    # Header row structure.
    hu_header_row = _find_header_row(hu_ws, "Place Name")
    expected_headers = ["Place Name", "Tier", "Place FIPS"]
    for year in [2025, 2030, 2035]:
        expected_headers.extend([
            f"{year} Share-Trend",
            f"{year} HU Method",
            f"{year} Diff",
            f"{year} % Diff",
        ])
    actual_headers = [
        hu_ws.cell(row=hu_header_row, column=col_idx).value
        for col_idx in range(1, len(expected_headers) + 1)
    ]
    assert actual_headers == expected_headers

    # Count data rows (should be 18: 9 HIGH + 9 MODERATE).
    data_rows = []
    row = hu_header_row + 1
    while hu_ws.cell(row=row, column=1).value:
        data_rows.append(row)
        row += 1
    assert len(data_rows) == 18

    # Validate HIGH places come before MODERATE places.
    tiers = [hu_ws.cell(row=r, column=2).value for r in data_rows]
    high_count = sum(1 for t in tiers if t == "HIGH")
    moderate_count = sum(1 for t in tiers if t == "MODERATE")
    assert high_count == 9
    assert moderate_count == 9
    # All HIGH rows should come before all MODERATE rows.
    first_moderate_idx = tiers.index("MODERATE")
    assert all(t == "HIGH" for t in tiers[:first_moderate_idx])
    assert all(t == "MODERATE" for t in tiers[first_moderate_idx:])

    # Verify numeric values are present and Diff = HU - Share.
    first_data_row = data_rows[0]
    share_2025 = hu_ws.cell(row=first_data_row, column=4).value
    hu_2025 = hu_ws.cell(row=first_data_row, column=5).value
    diff_2025 = hu_ws.cell(row=first_data_row, column=6).value
    pct_2025 = hu_ws.cell(row=first_data_row, column=7).value

    assert isinstance(share_2025, (int, float))
    assert isinstance(hu_2025, (int, float))
    assert isinstance(diff_2025, (int, float))
    assert isinstance(pct_2025, (int, float))
    assert abs(diff_2025 - (hu_2025 - share_2025)) < 0.01
    if share_2025 != 0:
        expected_pct = (hu_2025 - share_2025) / share_2025
        assert abs(pct_2025 - expected_pct) < 0.001


def test_hu_comparison_absent_without_data(monkeypatch, tmp_path: Path) -> None:
    """Without HU projection data, workbook should have 21 sheets (no HU Comparison)."""
    scenario = "baseline"
    _write_county_reference_csv(tmp_path)
    _write_synthetic_place_outputs(tmp_path, scenario=scenario)

    config = {
        "pipeline": {
            "projection": {"output_dir": "data/projections"},
            "export": {"output_dir": "data/exports"},
        },
        "place_projections": {"output": {"key_years": KEY_YEARS}},
    }

    monkeypatch.setattr(workbook_mod, "project_root", tmp_path)

    output_path = workbook_mod.build_workbook(
        scenario=scenario,
        config=config,
        date_stamp="20260301",
    )

    wb = load_workbook(output_path, data_only=True)
    assert len(wb.sheetnames) == 21
    assert "HU Comparison" not in wb.sheetnames
