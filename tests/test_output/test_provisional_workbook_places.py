"""Contract tests for PP-003 IMP-15 provisional workbook place-sheet wiring."""

from __future__ import annotations

import importlib
import sys
from datetime import date
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import load_workbook

PROJECT_ROOT = Path(__file__).resolve().parents[2]
sys.path.insert(0, str(PROJECT_ROOT))

workbook_mod = importlib.import_module("scripts.exports.build_provisional_workbook")

KEY_YEARS = workbook_mod.KEY_YEARS
SCENARIO_FACTORS = {
    "baseline": 1.0,
    "restricted_growth": 0.9,
    "high_growth": 1.1,
}


def _write_county_outputs(root: Path, scenario: str, factor: float) -> None:
    """Write minimal county projections used by the provisional workbook builder."""
    county_dir = root / "data" / "projections" / scenario / "county"
    county_dir.mkdir(parents=True, exist_ok=True)

    county_rows: list[dict[str, float | int]] = []
    for year_idx, year in enumerate(KEY_YEARS):
        yearly_base = 3000.0 + year_idx * 45.0
        for age in [0, 25, 50, 75]:
            county_rows.append(
                {
                    "year": year,
                    "age": age,
                    "population": (yearly_base + age) * factor,
                }
            )

    county_df = pd.DataFrame(county_rows)
    county_path = county_dir / f"nd_county_38001_projection_2025_2055_{scenario}.parquet"
    county_df.to_parquet(county_path, index=False)

    yearly_totals = county_df.groupby("year")["population"].sum()
    base_population = float(yearly_totals.loc[workbook_mod.BASE_YEAR])
    final_population = float(yearly_totals.loc[workbook_mod.FINAL_YEAR])
    absolute_growth = final_population - base_population
    growth_rate = absolute_growth / base_population if base_population else 0.0

    summary_df = pd.DataFrame(
        [
            {
                "fips": 38001,
                "name": "Adams County",
                "base_population": base_population,
                "final_population": final_population,
                "absolute_growth": absolute_growth,
                "growth_rate": growth_rate,
            }
        ]
    )
    summary_df.to_csv(county_dir / "countys_summary.csv", index=False)


def _write_place_outputs(root: Path, scenario: str, factor: float) -> None:
    """Write minimal place outputs (including balance rows) for one scenario."""
    place_dir = root / "data" / "projections" / scenario / "place"
    qa_dir = place_dir / "qa"
    qa_dir.mkdir(parents=True, exist_ok=True)

    place_values = {
        "3810001": [1000.0, 1080.0, 1160.0, 1240.0, 1320.0, 1400.0, 1480.0],
        "3810002": [500.0, 530.0, 560.0, 590.0, 620.0, 650.0, 680.0],
    }
    place_names = {"3810001": "Alpha City", "3810002": "Beta City"}
    place_tiers = {"3810001": "HIGH", "3810002": "LOWER"}

    summary_rows: list[dict[str, Any]] = []
    for place_fips, values in place_values.items():
        scaled_values = [value * factor for value in values]
        pd.DataFrame(
            {
                "year": KEY_YEARS,
                "population": scaled_values,
            }
        ).to_parquet(
            place_dir / f"nd_place_{place_fips}_projection_2025_2055_{scenario}.parquet",
            index=False,
        )

        base_population = scaled_values[0]
        final_population = scaled_values[-1]
        absolute_growth = final_population - base_population
        growth_rate = absolute_growth / base_population if base_population else 0.0
        summary_rows.append(
            {
                "place_fips": place_fips,
                "name": place_names[place_fips],
                "county_fips": "38001",
                "level": "place",
                "row_type": "place",
                "confidence_tier": place_tiers[place_fips],
                "base_population": base_population,
                "final_population": final_population,
                "absolute_growth": absolute_growth,
                "growth_rate": growth_rate,
                "base_share": 0.2,
                "final_share": 0.2,
                "processing_time": 0.01,
            }
        )

    balance_values = [200.0, 210.0, 220.0, 230.0, 240.0, 250.0, 260.0]
    scaled_balance = [value * factor for value in balance_values]
    summary_rows.append(
        {
            "place_fips": "bal_38001",
            "name": "Balance of Adams County",
            "county_fips": "38001",
            "level": "place",
            "row_type": "balance_of_county",
            "confidence_tier": None,
            "base_population": scaled_balance[0],
            "final_population": scaled_balance[-1],
            "absolute_growth": scaled_balance[-1] - scaled_balance[0],
            "growth_rate": (scaled_balance[-1] - scaled_balance[0]) / scaled_balance[0],
            "base_share": 0.1,
            "final_share": 0.1,
            "processing_time": 0.0,
        }
    )

    pd.DataFrame(summary_rows).to_csv(place_dir / "places_summary.csv", index=False)

    qa_rows = []
    for year, balance in zip(KEY_YEARS, scaled_balance, strict=False):
        place_total = sum(place_values[p][KEY_YEARS.index(year)] * factor for p in place_values)
        qa_rows.append(
            {
                "county_fips": "38001",
                "county_name": "Adams County",
                "year": year,
                "county_total": place_total + balance,
                "sum_of_places": place_total,
                "balance_of_county": balance,
                "balance_share": balance / (place_total + balance),
            }
        )
    pd.DataFrame(qa_rows).to_csv(qa_dir / "qa_balance_of_county.csv", index=False)


def _find_header_row(ws: Any, header_value: str) -> int:
    """Return row index containing expected first-column header value."""
    for row_idx in range(1, 80):
        if ws.cell(row=row_idx, column=1).value == header_value:
            return row_idx
    raise AssertionError(f"Header '{header_value}' not found in worksheet {ws.title}")


def test_provisional_workbook_places_sheet_contract(monkeypatch, tmp_path: Path) -> None:
    """Verify IMP-15 place sheets are added without breaking legacy sheets."""
    for scenario, factor in SCENARIO_FACTORS.items():
        _write_county_outputs(tmp_path, scenario=scenario, factor=factor)
        _write_place_outputs(tmp_path, scenario=scenario, factor=factor)

    monkeypatch.setattr(workbook_mod, "PROJECT_ROOT", tmp_path)
    monkeypatch.setattr(workbook_mod, "DATE_STAMP", "20260301")
    monkeypatch.setattr(workbook_mod, "TODAY", date(2026, 3, 1))

    rc = workbook_mod.main()
    assert rc == 0

    output_path = tmp_path / "data" / "exports" / "nd_population_projections_provisional_20260301.xlsx"
    assert output_path.exists()

    wb = load_workbook(output_path, data_only=True)
    expected_legacy_sheets = {
        "Table of Contents",
        "State Summary",
        "State Detail",
        "Scenario Comparison",
        "Counties — Baseline",
        "Counties — Restricted Growth",
        "Counties — High Growth",
        "Growth Rankings",
        "Age Structure",
    }
    assert expected_legacy_sheets.issubset(set(wb.sheetnames))
    assert sum(name.startswith("Counties — ") for name in wb.sheetnames) == 3

    for _scenario_key, scenario_short in workbook_mod.SCENARIO_SHORT_NAMES.items():
        sheet_name = f"Places — {scenario_short}"
        assert sheet_name in wb.sheetnames
        ws = wb[sheet_name]

        header_row = _find_header_row(ws, "Place FIPS")
        expected_headers = (
            ["Place FIPS", "Place", "County", "Row Type", "Tier"]
            + [str(year) for year in KEY_YEARS]
            + [
                f"Change ({workbook_mod.BASE_YEAR}–{workbook_mod.FINAL_YEAR})",
                f"% Change ({workbook_mod.BASE_YEAR}–{workbook_mod.FINAL_YEAR})",
            ]
        )
        actual_headers = [
            ws.cell(row=header_row, column=col_idx).value for col_idx in range(1, len(expected_headers) + 1)
        ]
        assert actual_headers == expected_headers

        data_rows: list[int] = []
        row_idx = header_row + 1
        while ws.cell(row=row_idx, column=1).value:
            data_rows.append(row_idx)
            row_idx += 1
        assert len(data_rows) == 3  # 2 places + 1 balance row

        row_types = [ws.cell(row=row_idx, column=4).value for row_idx in data_rows]
        assert "place" in row_types
        assert "balance_of_county" in row_types

        balance_row_idx = next(row_idx for row_idx in data_rows if ws.cell(row=row_idx, column=4).value == "balance_of_county")
        assert ws.cell(row=balance_row_idx, column=5).value == "BALANCE"
        for year_offset in range(len(KEY_YEARS)):
            assert ws.cell(row=balance_row_idx, column=6 + year_offset).value not in (None, "")

        # Growth column remains percent-formatted values derived from growth_rate.
        for row_idx in data_rows:
            assert ws.cell(row=row_idx, column=7 + len(KEY_YEARS)).value is not None
