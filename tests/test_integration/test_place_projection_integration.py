"""Integration coverage for PP-003 IMP-18 place projection pipeline flow."""

from __future__ import annotations

import importlib
import json
import sys
from pathlib import Path
from typing import Any

import numpy as np
import pandas as pd
import pytest
from openpyxl import load_workbook

from cohort_projections.data.process.place_shares import compute_historical_shares

PROJECT_ROOT = Path(__file__).parent.parent.parent
sys.path.insert(0, str(PROJECT_ROOT))

stage_mod = importlib.import_module("scripts.pipeline.02a_run_place_projections")
export_mod = importlib.import_module("scripts.pipeline.03_export_results")


def _synthetic_county_cohorts(years: list[int], totals_by_year: dict[int, float]) -> pd.DataFrame:
    """Build county projection rows with single-year age and sex detail."""
    rows: list[dict[str, object]] = []
    ages = list(range(91))
    sexes = ["Male", "Female"]
    cell_count = len(ages) * len(sexes)

    for year in years:
        cell_value = totals_by_year[year] / cell_count
        for age in ages:
            for sex in sexes:
                rows.append(
                    {
                        "year": year,
                        "age": age,
                        "sex": sex,
                        "race": "White alone, Non-Hispanic",
                        "population": cell_value,
                    }
                )
    return pd.DataFrame(rows)


def _write_backtest_winner(path: Path) -> Path:
    """Write synthetic winner payload consumed by the stage entrypoint."""
    payload = {
        "winner_variant_id": "B-II",
        "window": "primary",
        "score": 3.0757840903894844,
        "fitting_method": "wls",
        "constraint_method": "cap_and_redistribute",
        "acceptance": {"all_scored_tiers_pass_primary": True},
    }
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(payload, indent=2), encoding="utf-8")
    return path


def _write_crosswalk(path: Path) -> Path:
    """Write a projected-universe crosswalk (HIGH/MODERATE/LOWER in one county)."""
    crosswalk = pd.DataFrame(
        [
            {
                "state_fips": 38,
                "place_fips": 3825700,
                "place_name": "Fargo city",
                "county_fips": 38017,
                "assignment_type": "single_county",
                "area_share": 1.0,
                "historical_only": False,
                "source_vintage": 2020,
                "source_method": "test",
                "confidence_tier": "HIGH",
                "tier_boundary": False,
            },
            {
                "state_fips": 38,
                "place_fips": 3884780,
                "place_name": "Valley city",
                "county_fips": 38017,
                "assignment_type": "single_county",
                "area_share": 1.0,
                "historical_only": False,
                "source_vintage": 2020,
                "source_method": "test",
                "confidence_tier": "MODERATE",
                "tier_boundary": False,
            },
            {
                "state_fips": 38,
                "place_fips": 3890000,
                "place_name": "Smallville city",
                "county_fips": 38017,
                "assignment_type": "single_county",
                "area_share": 1.0,
                "historical_only": False,
                "source_vintage": 2020,
                "source_method": "test",
                "confidence_tier": "LOWER",
                "tier_boundary": False,
            },
        ]
    )
    path.parent.mkdir(parents=True, exist_ok=True)
    crosswalk.to_csv(path, index=False)
    return path


def _write_share_history(path: Path) -> Path:
    """Compute synthetic share-history parquet via the IMP-03 module."""
    county_history = {
        2020: 900.0,
        2021: 940.0,
        2022: 980.0,
        2023: 1020.0,
        2024: 1060.0,
    }
    share_patterns = {
        "3825700": {2020: 0.40, 2021: 0.41, 2022: 0.42, 2023: 0.43, 2024: 0.44},
        "3884780": {2020: 0.12, 2021: 0.12, 2022: 0.11, 2023: 0.11, 2024: 0.10},
        "3890000": {2020: 0.03, 2021: 0.029, 2022: 0.028, 2023: 0.027, 2024: 0.026},
    }
    names = {
        "3825700": "Fargo city",
        "3884780": "Valley city",
        "3890000": "Smallville city",
    }

    place_rows: list[dict[str, object]] = []
    for place_fips, yearly_shares in share_patterns.items():
        for year, share in yearly_shares.items():
            place_rows.append(
                {
                    "place_fips": place_fips,
                    "place_name": names[place_fips],
                    "county_fips": "38017",
                    "year": year,
                    "population": share * county_history[year],
                }
            )

    place_history = pd.DataFrame(place_rows)
    county_history_df = pd.DataFrame(
        {
            "county_fips": ["38017"] * len(county_history),
            "year": list(county_history.keys()),
            "county_population": list(county_history.values()),
        }
    )
    shares = compute_historical_shares(
        place_population_history=place_history,
        county_population_history=county_history_df,
        epsilon=0.001,
        year_min=2020,
        year_max=2024,
    )
    path.parent.mkdir(parents=True, exist_ok=True)
    shares.to_parquet(path, index=False)
    return path


def test_place_pipeline_stage_and_export_end_to_end(
    monkeypatch: pytest.MonkeyPatch,
    tmp_path: Path,
) -> None:
    """
    Run a synthetic subset end-to-end flow from share assembly to place export.

    Validates IMP-18 core contract:
    - stage entrypoint runs with real orchestrator
    - place hard constraints hold through QA artifacts
    - export `--places` produces summary + workbook outputs
    """
    projection_root = tmp_path / "projections"
    export_root = tmp_path / "exports"

    county_dir = projection_root / "baseline" / "county"
    county_dir.mkdir(parents=True, exist_ok=True)
    county_projection = _synthetic_county_cohorts(
        years=[2025, 2026],
        totals_by_year={2025: 1100.0, 2026: 1200.0},
    )
    county_projection.to_parquet(
        county_dir / "nd_county_38017_projection_2025_2026_baseline.parquet",
        index=False,
    )

    crosswalk_path = _write_crosswalk(tmp_path / "data" / "processed" / "geographic" / "crosswalk.csv")
    shares_path = _write_share_history(tmp_path / "data" / "processed" / "place_shares.parquet")

    counties_path = tmp_path / "data" / "raw" / "geographic" / "nd_counties.csv"
    counties_path.parent.mkdir(parents=True, exist_ok=True)
    pd.DataFrame(
        [{"county_fips": 38017, "state_fips": 38, "county_name": "Cass County"}]
    ).to_csv(counties_path, index=False)

    winner_path = _write_backtest_winner(
        tmp_path / "data" / "backtesting" / "place_backtest_results" / "backtest_winner.json"
    )

    config: dict[str, Any] = {
        "project": {"base_year": 2025, "projection_horizon": 1},
        "output": {"compression": "gzip"},
        "demographics": {
            "race_ethnicity": {
                "categories": [
                    "White alone, Non-Hispanic",
                    "Black alone, Non-Hispanic",
                    "AIAN alone, Non-Hispanic",
                    "Asian/PI alone, Non-Hispanic",
                    "Two or more races, Non-Hispanic",
                    "Hispanic (any race)",
                ]
            }
        },
        "pipeline": {
            "projection": {"output_dir": str(projection_root)},
            "export": {
                "output_dir": str(export_root),
                "create_packages": False,
                "formats": ["csv"],
                "summaries": ["total_population_by_year"],
            },
        },
        "geography": {"reference_data": {"counties_file": str(counties_path)}},
        "scenarios": {"baseline": {"active": True}},
        "place_projections": {
            "enabled": True,
            "crosswalk_path": str(crosswalk_path),
            "historical_shares_path": str(shares_path),
            "model": {
                "epsilon": 0.001,
                "lambda_decay": 0.9,
                "reconciliation_flag_threshold": 0.05,
                "history_start": 2020,
                "history_end": 2024,
            },
            "output": {"base_year": 2025, "end_year": 2026, "key_years": [2025, 2026]},
        },
    }

    monkeypatch.setattr(stage_mod, "load_projection_config", lambda _: config)
    monkeypatch.setattr(
        sys,
        "argv",
        [
            "02a_run_place_projections.py",
            "--scenarios",
            "baseline",
            "--winner-file",
            str(winner_path),
        ],
    )
    assert stage_mod.main() == 0

    place_dir = projection_root / "baseline" / "place"
    qa_dir = place_dir / "qa"
    assert place_dir.exists()
    assert (place_dir / "places_summary.csv").exists()
    assert (place_dir / "places_metadata.json").exists()
    for artifact_name in [
        "qa_tier_summary.csv",
        "qa_share_sum_validation.csv",
        "qa_reconciliation_magnitude.csv",
        "qa_outlier_flags.csv",
        "qa_balance_of_county.csv",
    ]:
        assert (qa_dir / artifact_name).exists()

    share_sum_df = pd.read_csv(qa_dir / "qa_share_sum_validation.csv")
    assert share_sum_df["constraint_satisfied"].all()
    np.testing.assert_allclose(
        share_sum_df["sum_place_shares"] + share_sum_df["balance_of_county_share"],
        1.0,
        rtol=1e-9,
        atol=1e-9,
    )

    reconciliation_df = pd.read_csv(qa_dir / "qa_reconciliation_magnitude.csv")
    assert (reconciliation_df["reconciliation_adjustment"] >= 0).all()
    np.testing.assert_allclose(
        reconciliation_df["total_after_adjustment"],
        1.0,
        rtol=1e-9,
        atol=1e-9,
    )

    balance_df = pd.read_csv(qa_dir / "qa_balance_of_county.csv")
    np.testing.assert_allclose(
        balance_df["sum_of_places"] + balance_df["balance_of_county"],
        balance_df["county_total"],
        rtol=1e-9,
        atol=1e-9,
    )

    summary_df = pd.read_csv(place_dir / "places_summary.csv")
    assert len(summary_df[summary_df["row_type"] == "place"]) == 3
    assert len(summary_df[summary_df["row_type"] == "balance_of_county"]) == 1

    monkeypatch.setattr(export_mod, "load_projection_config", lambda _: config)
    monkeypatch.setattr(
        sys,
        "argv",
        [
            "03_export_results.py",
            "--places",
            "--scenarios",
            "baseline",
            "--formats",
            "csv",
            "--no-package",
        ],
    )
    assert export_mod.main() == 0

    exported_summary = export_root / "baseline" / "place" / "places_summary.csv"
    assert exported_summary.exists()
    workbook_paths = sorted(export_root.glob("nd_projections_baseline_places_*.xlsx"))
    assert len(workbook_paths) == 1
    workbook = load_workbook(workbook_paths[0], read_only=True, data_only=True)
    assert "Methodology" in workbook.sheetnames
    workbook.close()

    converted_csv = sorted((export_root / "baseline" / "place" / "csv").glob("*.csv*"))
    assert len(converted_csv) >= 3
    assert (export_root / "data_dictionary.json").exists()
    assert (export_root / "data_dictionary.md").exists()
