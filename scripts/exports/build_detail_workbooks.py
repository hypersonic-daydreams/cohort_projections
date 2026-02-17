#!/usr/bin/env python3
"""
Build per-scenario detail workbooks for ND population projections.

Creates one workbook per active scenario, each containing age-sex detail
for the state, 8 planning regions, and all 53 counties. Designed to match
and improve upon the SDC 2024 County Projections format (ADR-038).

Output: data/exports/nd_projections_{scenario}_detail_{datestamp}.xlsx

Usage:
    python scripts/exports/build_detail_workbooks.py
    python scripts/exports/build_detail_workbooks.py --scenarios baseline restricted_growth
"""

import argparse
import sys
from datetime import UTC, datetime
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

PROJECT_ROOT = Path(__file__).parent.parent.parent
TODAY = datetime.now(tz=UTC).date()
DATE_STAMP = TODAY.strftime("%Y%m%d")
BASE_YEAR = 2025
FINAL_YEAR = 2045
KEY_YEARS = [2025, 2030, 2035, 2040, 2045]
PROVISIONAL_LABEL = "PROVISIONAL \u2014 Pending Review \u2014 Subject to Change"

SCENARIOS = {
    "baseline": "Baseline (Trend Continuation)",
    "restricted_growth": "Restricted Growth (CBO Policy-Adjusted)",
    "high_growth": "High Growth (Pre-Policy Elevated Immigration)",
}

AGE_GROUP_BINS = list(range(0, 90, 5)) + [91]  # [0,5,...,85,91]
AGE_GROUP_LABELS = [
    "0-4",
    "5-9",
    "10-14",
    "15-19",
    "20-24",
    "25-29",
    "30-34",
    "35-39",
    "40-44",
    "45-49",
    "50-54",
    "55-59",
    "60-64",
    "65-69",
    "70-74",
    "75-79",
    "80-84",
    "85+",
]

# Dependency ratio age group definitions
YOUTH_GROUPS = ["0-4", "5-9", "10-14"]
WORKING_GROUPS = [
    "15-19",
    "20-24",
    "25-29",
    "30-34",
    "35-39",
    "40-44",
    "45-49",
    "50-54",
    "55-59",
    "60-64",
]
AGED_GROUPS = ["65-69", "70-74", "75-79", "80-84", "85+"]

# ---------------------------------------------------------------------------
# ND Planning Regions (SDC / Human Service Zone conventions, ADR-038)
# ---------------------------------------------------------------------------
REGION_NAMES = {
    1: "Williston",
    2: "Minot",
    3: "Devils Lake",
    4: "Grand Forks",
    5: "Fargo",
    6: "Jamestown",
    7: "Bismarck",
    8: "Dickinson",
}

COUNTY_TO_REGION = {
    "38001": 8,
    "38003": 6,
    "38005": 3,
    "38007": 8,
    "38009": 2,
    "38011": 8,
    "38013": 2,
    "38015": 7,
    "38017": 5,
    "38019": 3,
    "38021": 6,
    "38023": 1,
    "38025": 8,
    "38027": 3,
    "38029": 7,
    "38031": 6,
    "38033": 8,
    "38035": 4,
    "38037": 7,
    "38039": 6,
    "38041": 8,
    "38043": 7,
    "38045": 6,
    "38047": 6,
    "38049": 2,
    "38051": 6,
    "38053": 1,
    "38055": 7,
    "38057": 7,
    "38059": 7,
    "38061": 2,
    "38063": 4,
    "38065": 7,
    "38067": 4,
    "38069": 2,
    "38071": 3,
    "38073": 5,
    "38075": 2,
    "38077": 5,
    "38079": 3,
    "38081": 5,
    "38083": 7,
    "38085": 7,
    "38087": 8,
    "38089": 8,
    "38091": 5,
    "38093": 6,
    "38095": 3,
    "38097": 5,
    "38099": 4,
    "38101": 2,
    "38103": 6,
    "38105": 1,
}

# Invert: region number -> sorted list of county FIPS
REGION_COUNTIES: dict[int, list[str]] = {}
for _fips, _reg in COUNTY_TO_REGION.items():
    REGION_COUNTIES.setdefault(_reg, []).append(_fips)
for _reg in REGION_COUNTIES:
    REGION_COUNTIES[_reg].sort()

# ---------------------------------------------------------------------------
# Style constants
# ---------------------------------------------------------------------------
HEADER_FONT = Font(name="Aptos", size=14, bold=True, color="1F3864")
SUBTITLE_FONT = Font(name="Aptos", size=11, italic=True, color="595959")
PROVISIONAL_FONT = Font(name="Aptos", size=11, bold=True, color="C00000")
SECTION_FONT = Font(name="Aptos", size=11, bold=True, color="1F3864")
COL_HEADER_FONT = Font(name="Aptos", size=10, bold=True, color="FFFFFF")
COL_HEADER_FILL = PatternFill("solid", fgColor="1F3864")
COL_HEADER_ALIGN = Alignment(horizontal="center")
TOTAL_FONT = Font(name="Aptos", size=10, bold=True)
NORMAL_FONT = Font(name="Aptos", size=10)
LINK_FONT = Font(name="Aptos", size=10, color="0563C1", underline="single")
TOC_SECTION_FONT = Font(name="Aptos", size=11, bold=True, color="1F3864")
METHODOLOGY_FONT = Font(name="Aptos", size=9, italic=True, color="595959")
NUM_FMT = "#,##0"
CHANGE_FMT = "+#,##0;\u2212#,##0;0"
PCT_FMT = "+0.0%;\u22120.0%;0.0%"
RATIO_FMT = "0.00"
THIN_BORDER = Border(bottom=Side(style="thin", color="D9D9D9"))
TOTAL_BORDER = Border(
    top=Side(style="thin", color="1F3864"),
    bottom=Side(style="double", color="1F3864"),
)


# ===================================================================
# Data loading
# ===================================================================


def load_county_names() -> dict[str, str]:
    """Load county FIPS -> short name mapping from geographic reference CSV."""
    csv_path = PROJECT_ROOT / "data" / "raw" / "geographic" / "nd_counties.csv"
    df = pd.read_csv(csv_path, dtype=str)
    nd = df[df["state_fips"] == "38"]
    names: dict[str, str] = {}
    for _, row in nd.iterrows():
        fips = row["county_fips"].zfill(5)
        name = row["county_name"]
        name = name.removesuffix(" County")
        names[fips] = name
    return names


def load_scenario_data(scenario: str) -> pd.DataFrame:
    """Load all 53 county parquet files for a scenario.

    Returns a DataFrame with columns:
        county_fips, age_group (str), sex, year, population
    filtered to KEY_YEARS and aggregated into 5-year age groups.
    """
    county_dir = PROJECT_ROOT / "data" / "projections" / scenario / "county"
    parquet_files = sorted(county_dir.glob("nd_county_*.parquet"))
    if not parquet_files:
        raise FileNotFoundError(f"No parquet files found in {county_dir}")

    dfs = []
    for pf in parquet_files:
        parts = pf.stem.split("_")
        fips = parts[2]  # nd_county_38XXX_...
        df = pd.read_parquet(pf)
        df["county_fips"] = fips
        dfs.append(df)

    all_data = pd.concat(dfs, ignore_index=True)

    # Assign 5-year age groups
    all_data["age_group"] = pd.cut(
        all_data["age"],
        bins=AGE_GROUP_BINS,
        right=False,
        labels=AGE_GROUP_LABELS,
    )
    all_data["age_group"] = all_data["age_group"].astype(str)

    # Filter to key years
    all_data = all_data[all_data["year"].isin(KEY_YEARS)]

    # Aggregate across race (sum all races) by county × age_group × sex × year
    grouped = (
        all_data.groupby(["county_fips", "age_group", "sex", "year"], observed=True)["population"]
        .sum()
        .reset_index()
    )

    return grouped


# ===================================================================
# Table building
# ===================================================================


def build_tables(data: pd.DataFrame) -> dict[str, pd.DataFrame]:
    """Build age-sex pivot tables from grouped data.

    Returns dict with keys "Male", "Female", "Both", each mapping to
    a DataFrame with index=AGE_GROUP_LABELS and columns=KEY_YEARS.
    """
    tables: dict[str, pd.DataFrame] = {}

    for sex in ["Male", "Female"]:
        sex_data = data[data["sex"] == sex]
        if sex_data.empty:
            tables[sex] = pd.DataFrame(0.0, index=AGE_GROUP_LABELS, columns=KEY_YEARS)
            continue
        pivot = sex_data.pivot_table(
            index="age_group",
            columns="year",
            values="population",
            aggfunc="sum",
        )
        tables[sex] = pivot.reindex(AGE_GROUP_LABELS).reindex(columns=KEY_YEARS).fillna(0.0)

    # Both sexes
    if data.empty:
        tables["Both"] = pd.DataFrame(0.0, index=AGE_GROUP_LABELS, columns=KEY_YEARS)
    else:
        pivot = data.pivot_table(
            index="age_group",
            columns="year",
            values="population",
            aggfunc="sum",
        )
        tables["Both"] = pivot.reindex(AGE_GROUP_LABELS).reindex(columns=KEY_YEARS).fillna(0.0)

    return tables


# ===================================================================
# Sheet writing
# ===================================================================


def _write_sex_section(ws, start_row: int, section_label: str, table: pd.DataFrame) -> int:
    """Write one sex section (header row + 18 age groups + total).

    Returns the next available row.
    """
    row = start_row

    # Section header
    ws.cell(row=row, column=1, value=section_label).font = SECTION_FONT
    row += 1

    # Column headers
    headers = ["Age Group"] + [str(y) for y in KEY_YEARS] + ["Change", "% Change"]
    for col_idx, hdr in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col_idx, value=hdr)
        cell.font = COL_HEADER_FONT
        cell.fill = COL_HEADER_FILL
        cell.alignment = COL_HEADER_ALIGN
    row += 1

    # Data rows
    for ag in AGE_GROUP_LABELS:
        ws.cell(row=row, column=1, value=ag).font = NORMAL_FONT
        for ci, year in enumerate(KEY_YEARS, 2):
            val = float(table.at[ag, year]) if year in table.columns else 0.0  # type: ignore[arg-type]
            cell = ws.cell(row=row, column=ci, value=round(val))
            cell.font = NORMAL_FONT
            cell.number_format = NUM_FMT

        first_val = float(table.at[ag, KEY_YEARS[0]]) if KEY_YEARS[0] in table.columns else 0.0  # type: ignore[arg-type]
        last_val = float(table.at[ag, KEY_YEARS[-1]]) if KEY_YEARS[-1] in table.columns else 0.0  # type: ignore[arg-type]
        change = last_val - first_val
        pct = change / first_val if first_val > 0 else 0.0

        ws.cell(row=row, column=7, value=round(change)).font = NORMAL_FONT
        ws.cell(row=row, column=7).number_format = CHANGE_FMT
        ws.cell(row=row, column=8, value=pct).font = NORMAL_FONT
        ws.cell(row=row, column=8).number_format = PCT_FMT

        for ci in range(1, 9):
            ws.cell(row=row, column=ci).border = THIN_BORDER
        row += 1

    # Total row
    ws.cell(row=row, column=1, value="Total").font = TOTAL_FONT
    for ci, year in enumerate(KEY_YEARS, 2):
        total = float(table[year].sum()) if year in table.columns else 0.0
        cell = ws.cell(row=row, column=ci, value=round(total))
        cell.font = TOTAL_FONT
        cell.number_format = NUM_FMT

    t_first = float(table[KEY_YEARS[0]].sum()) if KEY_YEARS[0] in table.columns else 0.0
    t_last = float(table[KEY_YEARS[-1]].sum()) if KEY_YEARS[-1] in table.columns else 0.0
    t_change = t_last - t_first
    t_pct = t_change / t_first if t_first > 0 else 0.0

    ws.cell(row=row, column=7, value=round(t_change)).font = TOTAL_FONT
    ws.cell(row=row, column=7).number_format = CHANGE_FMT
    ws.cell(row=row, column=8, value=t_pct).font = TOTAL_FONT
    ws.cell(row=row, column=8).number_format = PCT_FMT

    for ci in range(1, 9):
        ws.cell(row=row, column=ci).border = TOTAL_BORDER
    row += 1

    return row


def _write_dependency_ratios(ws, start_row: int, both_table: pd.DataFrame) -> int:
    """Write dependency ratio section. Returns next available row."""
    row = start_row

    ws.cell(row=row, column=1, value="Key Indicators").font = SECTION_FONT
    row += 1

    # Headers
    headers = ["Indicator"] + [str(y) for y in KEY_YEARS]
    for ci, hdr in enumerate(headers, 1):
        cell = ws.cell(row=row, column=ci, value=hdr)
        cell.font = COL_HEADER_FONT
        cell.fill = COL_HEADER_FILL
        cell.alignment = COL_HEADER_ALIGN
    row += 1

    indicators = [
        ("Youth Dependency (0-14 / 15-64)", YOUTH_GROUPS, []),
        ("Aged Dependency (65+ / 15-64)", [], AGED_GROUPS),
        ("Total Dependency", YOUTH_GROUPS, AGED_GROUPS),
    ]
    for label, youth_groups, aged_groups in indicators:
        ws.cell(row=row, column=1, value=label).font = NORMAL_FONT
        for ci, year in enumerate(KEY_YEARS, 2):
            if year not in both_table.columns:
                continue
            youth = float(both_table.loc[youth_groups, year].sum()) if youth_groups else 0.0
            aged = float(both_table.loc[aged_groups, year].sum()) if aged_groups else 0.0
            working = float(both_table.loc[WORKING_GROUPS, year].sum())
            ratio = (youth + aged) / working if working > 0 else 0.0
            cell = ws.cell(row=row, column=ci, value=ratio)
            cell.font = NORMAL_FONT
            cell.number_format = RATIO_FMT
        row += 1

    return row


def write_geography_sheet(
    ws, geo_name: str, tables: dict[str, pd.DataFrame], scenario_label: str
) -> None:
    """Write a complete geography sheet with Male/Female/Both tables."""
    row = 1

    # Header block
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
    ws.cell(row=row, column=1, value=geo_name).font = HEADER_FONT
    row += 1

    ws.cell(row=row, column=1, value=f"Scenario: {scenario_label}").font = SUBTITLE_FONT
    row += 1

    ws.cell(row=row, column=1, value=PROVISIONAL_LABEL).font = PROVISIONAL_FONT
    row += 1
    row += 1  # blank

    # Three sex sections
    for sex_key, display_label in [("Male", "Male"), ("Female", "Female"), ("Both", "Both Sexes")]:
        row = _write_sex_section(ws, row, display_label, tables[sex_key])
        row += 1  # blank between sections

    # Dependency ratios
    row = _write_dependency_ratios(ws, row, tables["Both"])
    row += 1

    # Back-to-TOC link
    cell = ws.cell(row=row, column=1, value="\u2190 Back to Table of Contents")
    cell.font = LINK_FONT
    cell.hyperlink = "#'Table of Contents'!A1"

    # Column widths
    ws.column_dimensions["A"].width = 16
    for ci in range(2, 9):
        ws.column_dimensions[get_column_letter(ci)].width = 13

    # Print setup
    ws.page_setup.orientation = "landscape"


# ===================================================================
# Table of Contents
# ===================================================================


def build_toc(ws, geo_registry: list[dict], scenario_label: str) -> None:
    """Populate the Table of Contents sheet with summary + hyperlinks."""
    row = 1

    # Title block
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
    ws.cell(row=row, column=1, value="North Dakota Population Projections").font = HEADER_FONT
    row += 1

    ws.cell(row=row, column=1, value=f"Scenario: {scenario_label}").font = SUBTITLE_FONT
    row += 1

    ws.cell(row=row, column=1, value=PROVISIONAL_LABEL).font = PROVISIONAL_FONT
    row += 1

    ws.cell(
        row=row,
        column=1,
        value=f"Generated: {TODAY.strftime('%B %d, %Y')}  |  Base: Census PEP 2000-2024",
    ).font = METHODOLOGY_FONT
    row += 1
    row += 1  # blank

    # Column headers
    headers = ["Geography", "Region", "2025 Pop", "2045 Pop", "Change", "% Change"]
    for ci, hdr in enumerate(headers, 1):
        cell = ws.cell(row=row, column=ci, value=hdr)
        cell.font = COL_HEADER_FONT
        cell.fill = COL_HEADER_FILL
        cell.alignment = COL_HEADER_ALIGN
    row += 1

    ws.freeze_panes = ws.cell(row=row, column=1)

    # Helper to write one entry row
    def _write_entry(
        name: str, region_label: str, pop_2025: float, pop_2045: float, sheet_name: str
    ) -> None:
        nonlocal row
        cell = ws.cell(row=row, column=1, value=name)
        cell.font = LINK_FONT
        cell.hyperlink = f"#'{sheet_name}'!A1"

        ws.cell(row=row, column=2, value=region_label).font = NORMAL_FONT

        ws.cell(row=row, column=3, value=round(pop_2025)).font = NORMAL_FONT
        ws.cell(row=row, column=3).number_format = NUM_FMT

        ws.cell(row=row, column=4, value=round(pop_2045)).font = NORMAL_FONT
        ws.cell(row=row, column=4).number_format = NUM_FMT

        change = pop_2045 - pop_2025
        ws.cell(row=row, column=5, value=round(change)).font = NORMAL_FONT
        ws.cell(row=row, column=5).number_format = CHANGE_FMT

        pct = change / pop_2025 if pop_2025 > 0 else 0.0
        ws.cell(row=row, column=6, value=pct).font = NORMAL_FONT
        ws.cell(row=row, column=6).number_format = PCT_FMT

        for ci in range(1, 7):
            ws.cell(row=row, column=ci).border = THIN_BORDER
        row += 1

    # STATE section
    ws.cell(row=row, column=1, value="STATE").font = TOC_SECTION_FONT
    row += 1
    state_entries = [g for g in geo_registry if g["section"] == "state"]
    for entry in state_entries:
        _write_entry(
            entry["name"], "\u2014", entry["pop_2025"], entry["pop_2045"], entry["sheet_name"]
        )
    row += 1  # blank

    # REGIONS section
    ws.cell(row=row, column=1, value="PLANNING REGIONS").font = TOC_SECTION_FONT
    row += 1
    region_entries = [g for g in geo_registry if g["section"] == "region"]
    for entry in region_entries:
        _write_entry(
            entry["name"],
            entry["region"],
            entry["pop_2025"],
            entry["pop_2045"],
            entry["sheet_name"],
        )
    row += 1  # blank

    # COUNTIES section
    ws.cell(row=row, column=1, value="COUNTIES").font = TOC_SECTION_FONT
    row += 1
    county_entries = [g for g in geo_registry if g["section"] == "county"]
    for entry in county_entries:
        _write_entry(
            entry["name"],
            entry["region"],
            entry["pop_2025"],
            entry["pop_2045"],
            entry["sheet_name"],
        )
    row += 1  # blank

    # Methodology footer
    methodology_lines = [
        "Methodology:",
        "  Cohort-component model (single-year age, sex, 6 race/ethnicity categories)",
        f"  Base year: {BASE_YEAR} (Census PEP estimates)",
        f"  Projection horizon: {BASE_YEAR}\u2013{FINAL_YEAR} (annual steps; 5-year intervals shown)",
        "  Fertility: SEER age-specific rates (2018\u20132022 average)",
        "  Mortality: CDC life tables with 0.5% annual improvement",
        "  Migration: Census PEP components (2000\u20132024), convergence interpolation",
        "  Scenario adjustments grounded in CBO Demographic Outlook (ADR-037)",
    ]
    for line in methodology_lines:
        ws.cell(row=row, column=1, value=line).font = METHODOLOGY_FONT
        row += 1

    # Column widths
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 14
    for ci in range(3, 7):
        ws.column_dimensions[get_column_letter(ci)].width = 14


# ===================================================================
# Workbook builder
# ===================================================================


def build_workbook(scenario: str, scenario_label: str, county_names: dict[str, str]) -> Path:
    """Build one detail workbook for a single scenario.

    Returns the output file path.
    """
    print(f"\n{'=' * 60}")
    print(f"Building detail workbook: {scenario}")
    print(f"{'=' * 60}")

    # Load data
    print("  Loading projection data...")
    grouped = load_scenario_data(scenario)
    n_counties = grouped["county_fips"].nunique()
    print(f"  Loaded {len(grouped):,} records for {n_counties} counties")

    # Build per-county tables
    print("  Building county tables...")
    county_tables: dict[str, dict[str, pd.DataFrame]] = {}
    for fips in sorted(grouped["county_fips"].unique()):
        county_data = grouped[grouped["county_fips"] == fips]
        county_tables[fips] = build_tables(county_data)

    # Create workbook (TOC is first sheet)
    wb = Workbook()
    toc_ws = wb.active
    toc_ws.title = "Table of Contents"

    geo_registry: list[dict] = []

    # --- State sheet ---
    print("  Writing state sheet...")
    state_tables = build_tables(grouped)
    state_ws = wb.create_sheet("North Dakota")
    write_geography_sheet(state_ws, "North Dakota", state_tables, scenario_label)
    geo_registry.append(
        {
            "name": "North Dakota",
            "section": "state",
            "region": "",
            "sheet_name": "North Dakota",
            "pop_2025": state_tables["Both"][BASE_YEAR].sum(),
            "pop_2045": state_tables["Both"][FINAL_YEAR].sum(),
        }
    )

    # --- Region sheets ---
    print("  Writing region sheets...")
    for region_num in sorted(REGION_NAMES):
        fips_list = REGION_COUNTIES[region_num]
        region_data = grouped[grouped["county_fips"].isin(fips_list)]
        region_tables = build_tables(region_data)

        sheet_name = f"Reg {region_num} - {REGION_NAMES[region_num]}"
        full_name = f"Region {region_num} \u2014 {REGION_NAMES[region_num]}"
        ws = wb.create_sheet(sheet_name)
        write_geography_sheet(ws, full_name, region_tables, scenario_label)

        n = len(fips_list)
        geo_registry.append(
            {
                "name": full_name,
                "section": "region",
                "region": f"{n} counties",
                "sheet_name": sheet_name,
                "pop_2025": region_tables["Both"][BASE_YEAR].sum(),
                "pop_2045": region_tables["Both"][FINAL_YEAR].sum(),
            }
        )

    # --- County sheets ---
    print("  Writing county sheets...")
    for fips in sorted(county_tables):
        name = county_names.get(fips, fips)
        tables = county_tables[fips]
        ws = wb.create_sheet(name)
        write_geography_sheet(ws, f"{name} County", tables, scenario_label)

        region_num = COUNTY_TO_REGION.get(fips, 0)
        region_label = f"Reg {region_num}" if region_num else ""
        geo_registry.append(
            {
                "name": name,
                "section": "county",
                "region": region_label,
                "sheet_name": name,
                "pop_2025": tables["Both"][BASE_YEAR].sum(),
                "pop_2045": tables["Both"][FINAL_YEAR].sum(),
            }
        )

    # --- Populate TOC ---
    print("  Building table of contents...")
    build_toc(toc_ws, geo_registry, scenario_label)

    # Save
    output_dir = PROJECT_ROOT / "data" / "exports"
    output_dir.mkdir(parents=True, exist_ok=True)
    output_path = output_dir / f"nd_projections_{scenario}_detail_{DATE_STAMP}.xlsx"

    print(f"  Saving to {output_path}...")
    wb.save(str(output_path))
    file_size = output_path.stat().st_size / 1024
    print(f"  Done. File size: {file_size:.1f} KB, Sheets: {len(wb.sheetnames)}")

    return output_path


# ===================================================================
# CLI
# ===================================================================


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Build per-scenario detail workbooks for ND projections (ADR-038)",
    )
    parser.add_argument(
        "--scenarios",
        nargs="+",
        default=list(SCENARIOS.keys()),
        help=f"Scenarios to build (default: all). Options: {', '.join(SCENARIOS.keys())}",
    )
    return parser.parse_args()


def main() -> int:
    args = parse_args()

    print("=" * 60)
    print("ND Population Projections \u2014 Detail Workbook Builder")
    print(f"Date: {TODAY.isoformat()}")
    print(f"Scenarios: {', '.join(args.scenarios)}")
    print("=" * 60)

    # Load county names once
    county_names = load_county_names()
    print(f"Loaded {len(county_names)} county names")

    output_paths: list[Path] = []
    for scenario in args.scenarios:
        if scenario not in SCENARIOS:
            print(f"WARNING: Unknown scenario '{scenario}', skipping")
            continue
        path = build_workbook(scenario, SCENARIOS[scenario], county_names)
        output_paths.append(path)

    print(f"\n{'=' * 60}")
    print(f"Complete. Generated {len(output_paths)} workbooks:")
    for p in output_paths:
        print(f"  {p.name}")
    print("=" * 60)

    return 0


if __name__ == "__main__":
    sys.exit(main())
