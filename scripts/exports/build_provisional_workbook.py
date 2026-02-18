#!/usr/bin/env python3
"""
Build provisional Excel workbook for ND population projection results.

Creates a multi-sheet workbook with Excel tables, provisional labeling,
and a table of contents sheet. Designed for leadership review.

Usage:
    python scripts/exports/build_provisional_workbook.py
"""

import sys
from datetime import UTC, datetime
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Border, Font, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

sys.path.insert(0, str(Path(__file__).parent))
from _methodology import (
    CONDITIONAL_CAVEAT,
    DATA_AVAILABILITY_NOTE,
    METHODOLOGY_LINES,
    ORGANIZATION_ATTRIBUTION,
    PROVISIONAL_LABEL,
    SCENARIO_SHORT_NAMES,
    SCENARIOS,
)

PROJECT_ROOT = Path(__file__).parent.parent.parent
TODAY = datetime.now(tz=UTC).date()
DATE_STAMP = TODAY.strftime("%Y%m%d")
BASE_YEAR = 2025
FINAL_YEAR = 2055
KEY_YEARS = [2025, 2030, 2035, 2040, 2045, 2050, 2055]

# Style constants
HEADER_FONT = Font(name="Aptos", size=14, bold=True, color="1F3864")
SUBTITLE_FONT = Font(name="Aptos", size=11, italic=True, color="595959")
PROVISIONAL_FONT = Font(name="Aptos", size=11, bold=True, color="C00000")
TOC_LINK_FONT = Font(name="Aptos", size=11, color="0563C1", underline="single")
NORMAL_FONT = Font(name="Aptos", size=10)
BOLD_FONT = Font(name="Aptos", size=10, bold=True)
TABLE_STYLE = "TableStyleMedium2"
THIN_BORDER = Border(
    bottom=Side(style="thin", color="D9D9D9"),
)


def load_scenario_summaries() -> dict[str, pd.DataFrame]:
    """Load the county summary CSVs for each scenario."""
    summaries = {}
    for key, label in SCENARIO_SHORT_NAMES.items():
        path = PROJECT_ROOT / "data" / "projections" / key / "county" / "countys_summary.csv"
        if path.exists():
            df = pd.read_csv(path)
            df["scenario"] = label
            summaries[key] = df
    return summaries


def load_scenario_yearly_totals() -> dict[str, pd.DataFrame]:
    """Load parquet files and aggregate to county-year totals for each scenario."""
    results = {}
    for key in SCENARIOS:
        county_dir = PROJECT_ROOT / "data" / "projections" / key / "county"
        parquet_files = sorted(county_dir.glob("*_projection_*.parquet"))
        county_dfs = []
        for pf in parquet_files:
            fips = pf.stem.split("_")[2]  # nd_county_38XXX_...
            df = pd.read_parquet(pf)
            yearly = df.groupby("year")["population"].sum().reset_index()
            yearly["fips"] = int(fips)
            county_dfs.append(yearly)
        if county_dfs:
            results[key] = pd.concat(county_dfs, ignore_index=True)
    return results


def load_county_names() -> dict[int, str]:
    """Get county FIPS -> name mapping from the baseline summary."""
    path = PROJECT_ROOT / "data" / "projections" / "baseline" / "county" / "countys_summary.csv"
    df = pd.read_csv(path)
    return dict(zip(df["fips"], df["name"], strict=False))


def add_provisional_header(ws, title: str, start_row: int = 1) -> int:
    """Add a standardized provisional header block to a worksheet.

    Returns the next available row after the header.
    """
    ws.cell(row=start_row, column=1, value=title).font = HEADER_FONT
    ws.cell(row=start_row + 1, column=1, value=PROVISIONAL_LABEL).font = PROVISIONAL_FONT
    ws.cell(
        row=start_row + 2,
        column=1,
        value=f"Generated: {TODAY.strftime('%B %d, %Y')}",
    ).font = SUBTITLE_FONT
    return start_row + 4  # one blank row after header


def create_excel_table(
    ws,
    ref: str,
    name: str,
    style: str = TABLE_STYLE,
) -> None:
    """Create a formatted Excel table on the worksheet."""
    tab = Table(displayName=name, ref=ref)
    tab.tableStyleInfo = TableStyleInfo(
        name=style,
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    ws.add_table(tab)


def auto_width(ws, min_width: int = 10, max_width: int = 30) -> None:
    """Auto-fit column widths based on content."""
    for col in ws.columns:
        col_letter = get_column_letter(col[0].column)
        lengths = [len(str(cell.value)) for cell in col if cell.value is not None]
        if lengths:
            best = min(max(max(lengths) + 2, min_width), max_width)
            ws.column_dimensions[col_letter].width = best


def fmt_pop(val: float) -> int:
    """Round population to integer for display."""
    return int(round(val))


def fmt_pct(val: float) -> str:
    """Format a decimal as a percentage string."""
    return f"{val * 100:+.1f}%"


# ---------------------------------------------------------------------------
# Sheet builders
# ---------------------------------------------------------------------------


def build_toc(wb: Workbook, sheet_defs: list[dict]) -> None:
    """Sheet 1: Table of Contents."""
    ws = wb.active
    ws.title = "Table of Contents"
    ws.sheet_properties.tabColor = "1F3864"

    row = add_provisional_header(ws, "North Dakota Population Projections 2025–2045")

    ws.cell(row=row, column=1, value="Table of Contents").font = Font(
        name="Aptos", size=12, bold=True, color="1F3864"
    )
    row += 1

    # Write TOC table headers
    headers = ["Sheet #", "Sheet Name", "Description"]
    for c, h in enumerate(headers, 1):
        ws.cell(row=row, column=c, value=h)
    data_start = row
    row += 1

    for i, sd in enumerate(sheet_defs, 1):
        ws.cell(row=row, column=1, value=i)
        cell = ws.cell(row=row, column=2, value=sd["name"])
        cell.font = TOC_LINK_FONT
        # Internal hyperlink to sheet
        cell.hyperlink = f"#'{sd['name']}'!A1"
        ws.cell(row=row, column=3, value=sd["description"])
        row += 1

    # Create table
    table_ref = f"A{data_start}:C{row - 1}"
    create_excel_table(ws, table_ref, "TOC")

    row += 1
    # Methodology note
    ws.cell(row=row, column=1, value="Methodology Summary").font = Font(
        name="Aptos", size=12, bold=True, color="1F3864"
    )
    row += 1
    for line in METHODOLOGY_LINES:
        formatted = line.format(base_year=BASE_YEAR, final_year=FINAL_YEAR)
        ws.cell(row=row, column=1, value=formatted).font = NORMAL_FONT
        row += 1
    ws.cell(row=row, column=1, value=ORGANIZATION_ATTRIBUTION).font = NORMAL_FONT
    row += 1
    ws.cell(row=row, column=1, value=CONDITIONAL_CAVEAT).font = NORMAL_FONT
    row += 1
    ws.cell(row=row, column=1, value=DATA_AVAILABILITY_NOTE).font = NORMAL_FONT
    row += 1

    row += 1

    auto_width(ws, min_width=12, max_width=60)


def build_state_summary_wide(wb: Workbook, yearly: dict[str, pd.DataFrame]) -> None:
    """Sheet 2: Wide-format state summary — one row per year, scenarios side by side."""
    ws = wb.create_sheet("State Summary")
    ws.sheet_properties.tabColor = "2E75B6"

    row = add_provisional_header(ws, "State Population Summary by Scenario")

    # Compute state totals per scenario per year
    scenario_totals: dict[str, dict[int, float]] = {}
    for key, label in SCENARIO_SHORT_NAMES.items():
        if key not in yearly:
            continue
        df = yearly[key]
        state = df.groupby("year")["population"].sum()
        scenario_totals[label] = {int(yr): float(pop) for yr, pop in state.items()}  # type: ignore[call-overload]

    all_years = sorted(set().union(*(yt.keys() for yt in scenario_totals.values())))
    base_pops = {label: yt.get(BASE_YEAR, 0) for label, yt in scenario_totals.items()}

    headers = ["Year"]
    for label in SCENARIO_SHORT_NAMES.values():
        headers.append(f"{label}")
        headers.append(f"{label} % Chg from {BASE_YEAR}")
    for c, h in enumerate(headers, 1):
        ws.cell(row=row, column=c, value=h)
    data_start = row
    row += 1

    for yr in all_years:
        ws.cell(row=row, column=1, value=yr)
        col = 2
        for label in SCENARIO_SHORT_NAMES.values():
            pop = scenario_totals.get(label, {}).get(yr)
            if pop is not None:
                ws.cell(row=row, column=col, value=fmt_pop(pop)).number_format = "#,##0"
                bp = base_pops.get(label, 0)
                pct = round((pop - bp) / bp * 100, 2) if bp else 0
                ws.cell(row=row, column=col + 1, value=pct).number_format = '0.0"%"'
            col += 2
        row += 1

    table_ref = f"A{data_start}:{get_column_letter(len(headers))}{row - 1}"
    create_excel_table(ws, table_ref, "StateSummary")

    row += 1

    auto_width(ws)


def build_state_detail(wb: Workbook, yearly: dict[str, pd.DataFrame]) -> None:
    """State detail sheet: long format (scenario × year) for pivoting/charting."""
    ws = wb.create_sheet("State Detail")
    ws.sheet_properties.tabColor = "2E75B6"

    row = add_provisional_header(ws, "State Population Detail by Scenario and Year")

    # Build state-level yearly totals
    rows_data = []
    for key, label in SCENARIO_SHORT_NAMES.items():
        if key not in yearly:
            continue
        df = yearly[key]
        state = df.groupby("year")["population"].sum().reset_index()
        base_pop = float(state.loc[state["year"] == BASE_YEAR, "population"].iloc[0])
        for _, r in state.iterrows():
            yr = int(r["year"])
            pop = float(r["population"])
            change = pop - base_pop
            pct = change / base_pop if base_pop else 0
            rows_data.append(
                {
                    "Scenario": label,
                    "Year": yr,
                    "Population": fmt_pop(pop),
                    "Change from 2025": fmt_pop(change),
                    "% Change from 2025": round(pct * 100, 2),
                }
            )

    headers = list(rows_data[0].keys())
    for c, h in enumerate(headers, 1):
        ws.cell(row=row, column=c, value=h)
    data_start = row
    row += 1

    for rd in rows_data:
        for c, h in enumerate(headers, 1):
            cell = ws.cell(row=row, column=c, value=rd[h])
            if h == "% Change from 2025":
                cell.number_format = '0.0"%"'
            elif h in ("Population", "Change from 2025"):
                cell.number_format = "#,##0"
        row += 1

    table_ref = f"A{data_start}:{get_column_letter(len(headers))}{row - 1}"
    create_excel_table(ws, table_ref, "StateDetail")

    row += 1

    auto_width(ws)


def build_scenario_comparison(wb: Workbook, yearly: dict[str, pd.DataFrame]) -> None:
    """Sheet 3: Side-by-side scenario comparison at key years."""
    ws = wb.create_sheet("Scenario Comparison")
    ws.sheet_properties.tabColor = "2E75B6"

    row = add_provisional_header(ws, "Scenario Comparison — State Totals at Key Years")

    # Build wide format: Year | Baseline | Restricted Growth | High Growth | Spread
    headers = [
        "Year",
        "Baseline",
        "Restricted Growth",
        "High Growth",
        "Spread (High \u2212 Restricted)",
    ]
    for c, h in enumerate(headers, 1):
        ws.cell(row=row, column=c, value=h)
    data_start = row
    row += 1

    for yr in KEY_YEARS:
        vals = {}
        for key, label in SCENARIO_SHORT_NAMES.items():
            if key in yearly:
                df = yearly[key]
                state = df.groupby("year")["population"].sum()
                if yr in state.index:
                    vals[label] = fmt_pop(float(state.loc[yr]))
        ws.cell(row=row, column=1, value=yr)
        ws.cell(row=row, column=2, value=vals.get("Baseline", "")).number_format = "#,##0"
        ws.cell(row=row, column=3, value=vals.get("Restricted Growth", "")).number_format = "#,##0"
        ws.cell(row=row, column=4, value=vals.get("High Growth", "")).number_format = "#,##0"
        high = vals.get("High Growth", 0)
        low = vals.get("Restricted Growth", 0)
        if high and low:
            ws.cell(row=row, column=5, value=high - low).number_format = "#,##0"
        row += 1

    table_ref = f"A{data_start}:E{row - 1}"
    create_excel_table(ws, table_ref, "ScenarioComparison")

    row += 1

    auto_width(ws)


def build_county_detail(
    wb: Workbook,
    scenario_key: str,
    scenario_short: str,
    scenario_full: str,
    yearly: dict[str, pd.DataFrame],
    county_names: dict[int, str],
) -> None:
    """Build a county detail sheet for one scenario with population at key years."""
    sheet_name = f"Counties \u2014 {scenario_short}"
    ws = wb.create_sheet(sheet_name)

    colors = {"Baseline": "548235", "Restricted Growth": "C55A11", "High Growth": "BF8F00"}
    ws.sheet_properties.tabColor = colors.get(scenario_short, "808080")

    row = add_provisional_header(ws, f"County Projections \u2014 {scenario_full} Scenario")

    if scenario_key not in yearly:
        ws.cell(row=row, column=1, value="No data available for this scenario.")
        return

    df = yearly[scenario_key]

    # Pivot to county x year
    pivot = df.pivot_table(index="fips", columns="year", values="population", aggfunc="sum")

    # Build output rows
    headers = (
        ["FIPS", "County"]
        + [str(yr) for yr in KEY_YEARS]
        + [
            f"Change ({BASE_YEAR}\u2013{FINAL_YEAR})",
            f"% Change ({BASE_YEAR}\u2013{FINAL_YEAR})",
        ]
    )
    for c, h in enumerate(headers, 1):
        ws.cell(row=row, column=c, value=h)
    data_start = row
    row += 1

    county_rows = []
    for fips in sorted(pivot.index):
        name = county_names.get(fips, f"FIPS {fips}")
        base = float(pivot.loc[fips, BASE_YEAR]) if BASE_YEAR in pivot.columns else 0
        final = float(pivot.loc[fips, FINAL_YEAR]) if FINAL_YEAR in pivot.columns else 0
        change = final - base
        pct = (change / base * 100) if base else 0
        year_vals = {}
        for yr in KEY_YEARS:
            year_vals[yr] = fmt_pop(float(pivot.loc[fips, yr])) if yr in pivot.columns else ""
        county_rows.append((fips, name, year_vals, fmt_pop(change), round(pct, 1)))

    for fips, name, year_vals, change, pct in county_rows:
        ws.cell(row=row, column=1, value=fips)
        ws.cell(row=row, column=2, value=name)
        for i, yr in enumerate(KEY_YEARS):
            cell = ws.cell(row=row, column=3 + i, value=year_vals[yr])
            cell.number_format = "#,##0"
        ws.cell(row=row, column=3 + len(KEY_YEARS), value=change).number_format = "#,##0"
        ws.cell(row=row, column=4 + len(KEY_YEARS), value=pct).number_format = '0.0"%"'
        row += 1

    # State total row
    ws.cell(row=row, column=1, value="")
    ws.cell(row=row, column=2, value="STATE TOTAL").font = BOLD_FONT
    for i, yr in enumerate(KEY_YEARS):
        total = int(round(pivot[yr].sum())) if yr in pivot.columns else ""
        cell = ws.cell(row=row, column=3 + i, value=total)
        cell.number_format = "#,##0"
        cell.font = BOLD_FONT
    base_total = pivot[BASE_YEAR].sum() if BASE_YEAR in pivot.columns else 0
    final_total = pivot[FINAL_YEAR].sum() if FINAL_YEAR in pivot.columns else 0
    change_total = fmt_pop(final_total - base_total)
    pct_total = round((final_total - base_total) / base_total * 100, 1) if base_total else 0
    ws.cell(row=row, column=3 + len(KEY_YEARS), value=change_total).number_format = "#,##0"
    ws.cell(row=row, column=3 + len(KEY_YEARS)).font = BOLD_FONT
    ws.cell(row=row, column=4 + len(KEY_YEARS), value=pct_total).number_format = '0.0"%"'
    ws.cell(row=row, column=4 + len(KEY_YEARS)).font = BOLD_FONT
    row += 1

    table_ref = f"A{data_start}:{get_column_letter(len(headers))}{row - 1}"
    safe_name = scenario_short.replace(" ", "")
    create_excel_table(ws, table_ref, f"Counties{safe_name}")

    row += 1

    auto_width(ws)


def build_growth_rankings(
    wb: Workbook,
    summaries: dict[str, pd.DataFrame],
) -> None:
    """Sheet: County growth rankings for baseline scenario."""
    ws = wb.create_sheet("Growth Rankings")
    ws.sheet_properties.tabColor = "548235"

    row = add_provisional_header(ws, "County Growth Rankings — Baseline Scenario (2025–2045)")

    df = summaries.get("baseline")
    if df is None:
        ws.cell(row=row, column=1, value="No baseline data available.")
        return

    df = df.sort_values("growth_rate", ascending=False).copy()
    df["rank"] = range(1, len(df) + 1)
    df["growth_pct"] = df["growth_rate"] * 100

    headers = [
        "Rank",
        "FIPS",
        "County",
        f"Population {BASE_YEAR}",
        f"Population {FINAL_YEAR}",
        f"Change ({BASE_YEAR}-{FINAL_YEAR})",
        f"% Change ({BASE_YEAR}-{FINAL_YEAR})",
    ]
    for c, h in enumerate(headers, 1):
        ws.cell(row=row, column=c, value=h)
    data_start = row
    row += 1

    for _, r in df.iterrows():
        ws.cell(row=row, column=1, value=int(r["rank"]))
        ws.cell(row=row, column=2, value=int(r["fips"]))
        ws.cell(row=row, column=3, value=r["name"])
        ws.cell(row=row, column=4, value=fmt_pop(r["base_population"])).number_format = "#,##0"
        ws.cell(row=row, column=5, value=fmt_pop(r["final_population"])).number_format = "#,##0"
        ws.cell(row=row, column=6, value=fmt_pop(r["absolute_growth"])).number_format = "#,##0"
        ws.cell(row=row, column=7, value=round(r["growth_pct"], 1)).number_format = '0.0"%"'
        row += 1

    table_ref = f"A{data_start}:G{row - 1}"
    create_excel_table(ws, table_ref, "GrowthRankings")

    row += 1

    auto_width(ws)


def build_age_structure(wb: Workbook) -> None:
    """Sheet: Age group distribution for baseline at key years."""
    ws = wb.create_sheet("Age Structure")
    ws.sheet_properties.tabColor = "7030A0"

    row = add_provisional_header(ws, "Age Group Distribution — Baseline Scenario")

    # Load all baseline county parquets and aggregate to state
    county_dir = PROJECT_ROOT / "data" / "projections" / "baseline" / "county"
    parquet_files = list(county_dir.glob("*_projection_*.parquet"))
    all_data = pd.concat([pd.read_parquet(pf) for pf in parquet_files], ignore_index=True)

    # Define age groups
    age_bins = [0, 5, 18, 25, 45, 65, 200]
    age_labels = ["0–4", "5–17", "18–24", "25–44", "45–64", "65+"]
    all_data["age_group"] = pd.cut(all_data["age"], bins=age_bins, labels=age_labels, right=False)

    # Aggregate
    grouped = (
        all_data.groupby(["year", "age_group"], observed=True)["population"].sum().reset_index()
    )
    pivot = grouped.pivot_table(
        index="age_group", columns="year", values="population", aggfunc="sum"
    )

    headers = ["Age Group"] + [str(yr) for yr in KEY_YEARS] + ["Share 2025", "Share 2045"]
    for c, h in enumerate(headers, 1):
        ws.cell(row=row, column=c, value=h)
    data_start = row
    row += 1

    total_2025 = float(pivot[BASE_YEAR].sum()) if BASE_YEAR in pivot.columns else 1
    total_2045 = float(pivot[FINAL_YEAR].sum()) if FINAL_YEAR in pivot.columns else 1

    for ag in age_labels:
        ws.cell(row=row, column=1, value=ag)
        for i, yr in enumerate(KEY_YEARS):
            val = fmt_pop(float(pivot.loc[ag, yr])) if yr in pivot.columns else ""  # type: ignore[arg-type]
            ws.cell(row=row, column=2 + i, value=val).number_format = "#,##0"
        # Shares
        share_25 = (
            float(pivot.loc[ag, BASE_YEAR]) / total_2025 * 100  # type: ignore[arg-type]
            if BASE_YEAR in pivot.columns
            else 0
        )
        share_45 = (
            float(pivot.loc[ag, FINAL_YEAR]) / total_2045 * 100  # type: ignore[arg-type]
            if FINAL_YEAR in pivot.columns
            else 0
        )
        ws.cell(
            row=row, column=2 + len(KEY_YEARS), value=round(share_25, 1)
        ).number_format = '0.0"%"'
        ws.cell(
            row=row, column=3 + len(KEY_YEARS), value=round(share_45, 1)
        ).number_format = '0.0"%"'
        row += 1

    # Total row
    ws.cell(row=row, column=1, value="Total").font = BOLD_FONT
    for i, yr in enumerate(KEY_YEARS):
        val = fmt_pop(float(pivot[yr].sum())) if yr in pivot.columns else ""
        cell = ws.cell(row=row, column=2 + i, value=val)
        cell.number_format = "#,##0"
        cell.font = BOLD_FONT
    ws.cell(row=row, column=2 + len(KEY_YEARS), value=100.0).number_format = '0.0"%"'
    ws.cell(row=row, column=2 + len(KEY_YEARS)).font = BOLD_FONT
    ws.cell(row=row, column=3 + len(KEY_YEARS), value=100.0).number_format = '0.0"%"'
    ws.cell(row=row, column=3 + len(KEY_YEARS)).font = BOLD_FONT
    row += 1

    table_ref = f"A{data_start}:{get_column_letter(len(headers))}{row - 1}"
    create_excel_table(ws, table_ref, "AgeStructure")

    row += 1

    auto_width(ws)


def main() -> int:
    """Build the provisional workbook."""
    print("Loading projection data...")
    summaries = load_scenario_summaries()
    yearly = load_scenario_yearly_totals()
    county_names = load_county_names()

    print(f"Loaded {len(summaries)} scenario summaries")
    print(f"Loaded {len(yearly)} scenario yearly datasets")
    print(f"County names: {len(county_names)}")

    wb = Workbook()

    # Define sheet structure for TOC
    sheet_defs = [
        {
            "name": "State Summary",
            "description": "State-level population by scenario and year \u2014 side-by-side comparison",
        },
        {
            "name": "State Detail",
            "description": "State-level population \u2014 one row per scenario-year combination (for charts and pivot tables)",
        },
        {
            "name": "Scenario Comparison",
            "description": "Side-by-side comparison of all three scenarios at key milestone years",
        },
        {
            "name": "Counties \u2014 Baseline",
            "description": f"All 53 counties: projected population at key years under {SCENARIOS['baseline']} assumptions",
        },
        {
            "name": "Counties \u2014 Restricted Growth",
            "description": f"All 53 counties: projected population under {SCENARIOS['restricted_growth']} scenario (CBO time-varying migration, \u22125% fertility)",
        },
        {
            "name": "Counties \u2014 High Growth",
            "description": f"All 53 counties: projected population under {SCENARIOS['high_growth']} scenario (+15% migration, +5% fertility)",
        },
        {
            "name": "Growth Rankings",
            "description": f"Counties ranked by projected growth rate (baseline scenario, {BASE_YEAR}\u2013{FINAL_YEAR})",
        },
        {
            "name": "Age Structure",
            "description": "State-level age group distribution at key years (baseline scenario)",
        },
    ]

    print("Building Table of Contents...")
    build_toc(wb, sheet_defs)

    print("Building State Summary...")
    build_state_summary_wide(wb, yearly)

    print("Building State Detail...")
    build_state_detail(wb, yearly)

    print("Building Scenario Comparison...")
    build_scenario_comparison(wb, yearly)

    for key in SCENARIOS:
        short = SCENARIO_SHORT_NAMES[key]
        full = SCENARIOS[key]
        print(f"Building Counties \u2014 {short}...")
        build_county_detail(wb, key, short, full, yearly, county_names)

    print("Building Growth Rankings...")
    build_growth_rankings(wb, summaries)

    print("Building Age Structure...")
    build_age_structure(wb)

    # Save
    output_dir = PROJECT_ROOT / "data" / "exports"
    output_dir.mkdir(parents=True, exist_ok=True)
    output_path = output_dir / f"nd_population_projections_provisional_{DATE_STAMP}.xlsx"

    print(f"\nSaving to {output_path}...")
    wb.save(output_path)
    print(f"Done. File size: {output_path.stat().st_size / 1024:.1f} KB")
    return 0


if __name__ == "__main__":
    sys.exit(main())
