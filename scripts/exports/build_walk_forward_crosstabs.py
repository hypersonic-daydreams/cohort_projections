#!/usr/bin/env python3
"""Build walk-forward validation crosstab workbook.

Reads county-level and state-level annual error data produced by the
walk-forward validation script and pivots it into a crosstab layout
with origin-year rows and calendar-year columns, one sheet per method.

Input:
    data/analysis/walk_forward/annual_county_detail.csv
    data/analysis/walk_forward/annual_state_results.csv

Output:
    data/analysis/walk_forward/walk_forward_crosstabs.xlsx

Usage:
    python scripts/exports/build_walk_forward_crosstabs.py
"""

from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Alignment, Font, PatternFill, numbers
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

PROJECT_ROOT = Path(__file__).parent.parent.parent
DATA_DIR = PROJECT_ROOT / "data" / "analysis" / "walk_forward"

# Method code -> display label mapping
METHOD_LABELS = {
    "sdc_2024": "SDC 2024",
    "m2026": "M2026",
}

# Style constants (consistent with other export scripts)
HEADER_FONT = Font(name="Aptos", size=14, bold=True, color="1F3864")
SUBTITLE_FONT = Font(name="Aptos", size=11, italic=True, color="595959")
TABLE_STYLE = "TableStyleMedium2"

# Conditional formatting thresholds (absolute percent error)
GREEN_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
GREEN_FONT = Font(color="006100")
ORANGE_FILL = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
ORANGE_FONT = Font(color="9C5700")
RED_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
RED_FONT = Font(color="9C0006")


def load_county_detail() -> pd.DataFrame:
    """Load the annual county detail CSV from walk-forward validation."""
    path = DATA_DIR / "annual_county_detail.csv"
    df = pd.read_csv(path)
    return df


def load_state_results() -> pd.DataFrame:
    """Load the annual state-level results CSV from walk-forward validation."""
    path = DATA_DIR / "annual_state_results.csv"
    df = pd.read_csv(path)
    return df


def build_county_crosstab(df: pd.DataFrame, method: str) -> pd.DataFrame:
    """Pivot county detail into crosstab: rows = (county, origin_year), cols = validation_year.

    Values are percent error (pct_error column from the source data).
    """
    subset = df[df["method"] == method].copy()

    # Sort counties alphabetically, then by origin year within each county
    subset = subset.sort_values(["county_name", "origin_year", "validation_year"])

    # Build row label: "County Name (origin_year)"
    subset["row_label"] = (
        subset["county_name"] + " (" + subset["origin_year"].astype(str) + ")"
    )

    # Pivot: row_label x validation_year -> pct_error
    pivot = subset.pivot_table(
        index=["county_name", "origin_year", "row_label"],
        columns="validation_year",
        values="pct_error",
        aggfunc="first",
    )

    # Sort by county_name then origin_year (already done but ensure after pivot)
    pivot = pivot.sort_index(level=["county_name", "origin_year"])

    return pivot


def build_state_crosstab(df: pd.DataFrame, method: str) -> pd.DataFrame:
    """Pivot state results into crosstab: rows = origin_year, cols = validation_year.

    Values are pct_error.
    """
    subset = df[df["method"] == method].copy()
    subset = subset.sort_values(["origin_year", "validation_year"])
    subset["row_label"] = "North Dakota (" + subset["origin_year"].astype(str) + ")"

    pivot = subset.pivot_table(
        index=["origin_year", "row_label"],
        columns="validation_year",
        values="pct_error",
        aggfunc="first",
    )

    pivot = pivot.sort_index(level="origin_year")
    return pivot


def create_excel_table(
    ws,
    ref: str,
    name: str,
    style: str = TABLE_STYLE,
) -> None:
    """Create a formatted Excel table on the worksheet."""
    tab = Table(displayName=name, ref=ref)
    tab.tableStyleInfo = TableStyleInfo(
        name=style,
        showFirstColumn=True,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    ws.add_table(tab)


def auto_width(ws, min_width: int = 10, max_width: int = 30) -> None:
    """Auto-fit column widths based on content."""
    for col in ws.columns:
        col_letter = get_column_letter(col[0].column)
        lengths = [len(str(cell.value)) for cell in col if cell.value is not None]
        if lengths:
            best = min(max(max(lengths) + 2, min_width), max_width)
            ws.column_dimensions[col_letter].width = best


def apply_conditional_formatting(ws, data_start_row: int, data_end_row: int, data_start_col: int, data_end_col: int) -> None:
    """Apply three-tier conditional formatting to data cells.

    Thresholds (on absolute value of percent error):
        green:  |error| <= 2%
        orange: 2% < |error| <= 5%
        red:    |error| > 5%
    """
    start_col_letter = get_column_letter(data_start_col)
    end_col_letter = get_column_letter(data_end_col)
    cell_range = f"{start_col_letter}{data_start_row}:{end_col_letter}{data_end_row}"

    # Green: between -2 and 2
    ws.conditional_formatting.add(
        cell_range,
        CellIsRule(
            operator="between",
            formula=["-2", "2"],
            fill=GREEN_FILL,
            font=GREEN_FONT,
        ),
    )

    # Orange: > 2 and <= 5 (positive side)
    ws.conditional_formatting.add(
        cell_range,
        CellIsRule(
            operator="between",
            formula=["2.0001", "5"],
            fill=ORANGE_FILL,
            font=ORANGE_FONT,
        ),
    )

    # Orange: < -2 and >= -5 (negative side)
    ws.conditional_formatting.add(
        cell_range,
        CellIsRule(
            operator="between",
            formula=["-5", "-2.0001"],
            fill=ORANGE_FILL,
            font=ORANGE_FONT,
        ),
    )

    # Red: > 5
    ws.conditional_formatting.add(
        cell_range,
        CellIsRule(
            operator="greaterThan",
            formula=["5"],
            fill=RED_FILL,
            font=RED_FONT,
        ),
    )

    # Red: < -5
    ws.conditional_formatting.add(
        cell_range,
        CellIsRule(
            operator="lessThan",
            formula=["-5"],
            fill=RED_FILL,
            font=RED_FONT,
        ),
    )


def write_method_sheet(
    wb: Workbook,
    sheet_name: str,
    county_pivot: pd.DataFrame,
    state_pivot: pd.DataFrame,
    calendar_years: list[int],
) -> None:
    """Write a single method sheet with state + county rows and calendar-year columns."""
    ws = wb.create_sheet(title=sheet_name)

    # --- Header ---
    ws.cell(row=1, column=1, value=f"Walk-Forward Validation: {sheet_name}").font = HEADER_FONT
    ws.cell(
        row=2,
        column=1,
        value="Percent error by county, origin year, and calendar year. "
        "Green: |error| <= 2%. Orange: 2-5%. Red: > 5%.",
    ).font = SUBTITLE_FONT

    # Table starts at row 4
    table_start_row = 4

    # --- Column headers ---
    # Col A = "County (Origin Year)", then one column per calendar year
    # Excel Table headers must be strings, so convert year integers to strings
    ws.cell(row=table_start_row, column=1, value="County (Origin Year)")
    for j, year in enumerate(calendar_years):
        ws.cell(row=table_start_row, column=j + 2, value=str(year))

    # --- State rows first ---
    current_row = table_start_row + 1
    for (_origin_year, row_label), row_data in state_pivot.iterrows():
        ws.cell(row=current_row, column=1, value=row_label)
        # Bold state rows
        ws.cell(row=current_row, column=1).font = Font(name="Aptos", size=10, bold=True)
        for j, year in enumerate(calendar_years):
            if year in row_data.index and pd.notna(row_data[year]):
                cell = ws.cell(row=current_row, column=j + 2, value=round(row_data[year], 1))
                cell.number_format = '0.0"%"'
                cell.alignment = Alignment(horizontal="center")
        current_row += 1

    # --- County rows ---
    for (county_name, origin_year, row_label), row_data in county_pivot.iterrows():
        ws.cell(row=current_row, column=1, value=row_label)
        for j, year in enumerate(calendar_years):
            if year in row_data.index and pd.notna(row_data[year]):
                cell = ws.cell(row=current_row, column=j + 2, value=round(row_data[year], 1))
                cell.number_format = '0.0"%"'
                cell.alignment = Alignment(horizontal="center")
        current_row += 1

    # --- Create Excel Table ---
    data_end_row = current_row - 1
    num_cols = 1 + len(calendar_years)
    table_ref = f"A{table_start_row}:{get_column_letter(num_cols)}{data_end_row}"

    # Table names must be unique and contain no spaces
    table_name = sheet_name.replace(" ", "_") + "_Errors"
    create_excel_table(ws, table_ref, table_name)

    # --- Conditional formatting on data cells ---
    if data_end_row >= table_start_row + 1:
        apply_conditional_formatting(
            ws,
            data_start_row=table_start_row + 1,
            data_end_row=data_end_row,
            data_start_col=2,
            data_end_col=num_cols,
        )

    # --- Freeze panes: first column + header row ---
    ws.freeze_panes = "B5"

    # --- Column widths ---
    ws.column_dimensions["A"].width = 30  # County name column
    for j in range(len(calendar_years)):
        ws.column_dimensions[get_column_letter(j + 2)].width = 8


def write_method_sheet_by_year(
    wb: Workbook,
    sheet_name: str,
    county_pivot: pd.DataFrame,
    state_pivot: pd.DataFrame,
    calendar_years: list[int],
) -> None:
    """Write a transposed sheet: rows = calendar years, columns = runs (county × origin)."""
    ws = wb.create_sheet(title=sheet_name)

    # --- Header ---
    ws.cell(row=1, column=1, value=f"Walk-Forward Validation: {sheet_name}").font = HEADER_FONT
    ws.cell(
        row=2,
        column=1,
        value="Percent error by calendar year. Each column is one projection run. "
        "Green: |error| <= 2%. Orange: 2-5%. Red: > 5%.",
    ).font = SUBTITLE_FONT

    table_start_row = 4

    # Build column list: state runs first, then county runs
    col_labels: list[str] = []
    col_data: list[dict[int, float]] = []

    # State columns
    for (_origin_year, row_label), row_data in state_pivot.iterrows():
        col_labels.append(row_label)
        col_data.append(
            {int(yr): round(row_data[yr], 1) for yr in row_data.index if pd.notna(row_data[yr])}
        )

    # County columns
    for (_county_name, _origin_year, row_label), row_data in county_pivot.iterrows():
        col_labels.append(row_label)
        col_data.append(
            {int(yr): round(row_data[yr], 1) for yr in row_data.index if pd.notna(row_data[yr])}
        )

    # --- Column headers ---
    ws.cell(row=table_start_row, column=1, value="Year")
    for j, label in enumerate(col_labels):
        cell = ws.cell(row=table_start_row, column=j + 2, value=label)
        cell.alignment = Alignment(text_rotation=90, horizontal="center", vertical="bottom")
        # Bold state columns
        if label.startswith("North Dakota"):
            cell.font = Font(name="Aptos", size=10, bold=True)

    # --- Year rows ---
    current_row = table_start_row + 1
    for year in calendar_years:
        ws.cell(row=current_row, column=1, value=year)
        for j, data in enumerate(col_data):
            if year in data:
                cell = ws.cell(row=current_row, column=j + 2, value=data[year])
                cell.number_format = '0.0"%"'
                cell.alignment = Alignment(horizontal="center")
        current_row += 1

    # --- Create Excel Table ---
    data_end_row = current_row - 1
    num_cols = 1 + len(col_labels)
    table_ref = f"A{table_start_row}:{get_column_letter(num_cols)}{data_end_row}"

    table_name = sheet_name.replace(" ", "_").replace("(", "").replace(")", "") + "_ByYear"
    create_excel_table(ws, table_ref, table_name)

    # --- Conditional formatting ---
    if data_end_row >= table_start_row + 1:
        apply_conditional_formatting(
            ws,
            data_start_row=table_start_row + 1,
            data_end_row=data_end_row,
            data_start_col=2,
            data_end_col=num_cols,
        )

    # --- Freeze panes: year column + header row ---
    ws.freeze_panes = "B5"

    # --- Column widths ---
    ws.column_dimensions["A"].width = 8  # Year column
    for j in range(len(col_labels)):
        ws.column_dimensions[get_column_letter(j + 2)].width = 5

    # Row height for rotated headers
    ws.row_dimensions[table_start_row].height = 120


def main() -> None:
    """Build the walk-forward crosstabs workbook."""
    print("Loading walk-forward validation data...")
    county_df = load_county_detail()
    state_df = load_state_results()

    # Determine full range of calendar years across all data
    all_years = sorted(county_df["validation_year"].unique())
    print(f"  Calendar years: {all_years[0]}-{all_years[-1]}")
    print(f"  Methods: {list(county_df['method'].unique())}")
    print(f"  Origin years: {sorted(county_df['origin_year'].unique())}")
    print(f"  Counties: {county_df['county_name'].nunique()}")

    wb = Workbook()
    # Remove the default sheet (we create named sheets)
    wb.remove(wb.active)

    for method_code, method_label in METHOD_LABELS.items():
        print(f"\nBuilding sheet: {method_label}")

        county_pivot = build_county_crosstab(county_df, method_code)
        state_pivot = build_state_crosstab(state_df, method_code)

        n_county_rows = len(county_pivot)
        n_state_rows = len(state_pivot)
        print(f"  State rows: {n_state_rows}, County rows: {n_county_rows}")

        write_method_sheet(wb, method_label, county_pivot, state_pivot, all_years)

    # Pivoted sheets (years as rows, runs as columns)
    for method_code, method_label in METHOD_LABELS.items():
        pivoted_name = f"{method_label} by Year"
        print(f"\nBuilding sheet: {pivoted_name}")

        county_pivot = build_county_crosstab(county_df, method_code)
        state_pivot = build_state_crosstab(state_df, method_code)

        write_method_sheet_by_year(
            wb, pivoted_name, county_pivot, state_pivot, all_years
        )

    # Save
    output_path = DATA_DIR / "walk_forward_crosstabs.xlsx"
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(output_path))
    print(f"\nSaved: {output_path}")


if __name__ == "__main__":
    main()
