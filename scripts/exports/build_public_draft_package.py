#!/usr/bin/env python3
"""Build the PUB-2026 draft public marketing handoff package.

Produces a consolidated draft public Excel workbook, consolidated CSV download,
baseline population pyramids, and four reference key-chart PNGs from the latest projection
parquet exports under ``data/projections/baseline/``. Outputs are written into
the marketing handoff folder. After ADR-065, run this only after the baseline
projection outputs have been regenerated with CBO-adjusted assumptions.

Numbers reflect the locked-config production run (``m2026r1`` / ``cfg-20260611-production-lock``,
commit ``12fa6f9``, 2026-06-13; CF-001 disposed, ADR-067 applied). Outputs remain marked DRAFT
because they are a pre-publication marketing handoff, not because the numbers are provisional.
Rerun only if the locked production outputs change.

Usage:
    python scripts/exports/build_public_draft_package.py
    python scripts/exports/build_public_draft_package.py --skip-charts
"""

from __future__ import annotations

import argparse
import shutil
import sys
from datetime import UTC, datetime
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

PROJECT_ROOT = Path(__file__).parent.parent.parent
sys.path.insert(0, str(PROJECT_ROOT))
sys.path.insert(0, str(Path(__file__).parent))

from _methodology import (  # noqa: E402
    CONDITIONAL_CAVEAT,
    SCENARIOS,
)
from build_detail_workbooks import (  # noqa: E402
    COUNTY_TO_REGION,
    REGION_COUNTIES,
    REGION_NAMES,
    load_county_names,
)

TODAY = datetime.now(tz=UTC).date()
DATE_STAMP = TODAY.strftime("%Y%m%d")
BASE_YEAR = 2025
FINAL_YEAR = 2055
ALL_YEARS = list(range(BASE_YEAR, FINAL_YEAR + 1))
KEY_YEARS = [2025, 2030, 2035, 2040, 2045, 2050, 2055]

PUBLIC_SCENARIOS = ["baseline"]

# Output locations
HANDOFF_DIR = PROJECT_ROOT / "docs" / "plans" / "2026-public-projection-release-handoff"
MARKETING_DIR = HANDOFF_DIR / "marketing-ready"
DRAFTS_DIR = MARKETING_DIR / "drafts"

WORKBOOK_NAME = "PUB-2026 Draft Public Workbook.xlsx"
CSV_NAME = "PUB-2026 Draft Public Dataset.csv"
REFERENCE_PDF_NAME = "PUB-2026 Reference - 2024 SDC PDF.pdf"
STALE_ALTERNATIVE_SCENARIO_ARTIFACTS = [
    "chart_state_3scenario_line.png",
    "pyramid_state_2055_high.png",
    "pyramid_state_2055_restricted.png",
]

# Styling
HEADER_FONT = Font(name="Aptos", size=14, bold=True, color="1F3864")
SUBTITLE_FONT = Font(name="Aptos", size=11, italic=True, color="595959")
DRAFT_FONT = Font(name="Aptos", size=11, bold=True, color="C00000")
COL_HEADER_FONT = Font(name="Aptos", size=10, bold=True, color="FFFFFF")
COL_HEADER_FILL = PatternFill("solid", fgColor="1F3864")
COL_HEADER_ALIGN = Alignment(horizontal="center")
NORMAL_FONT = Font(name="Aptos", size=10)
SECTION_FONT = Font(name="Aptos", size=11, bold=True, color="1F3864")
NUM_FMT = "#,##0"
RATIO_FMT = "0.00"

DRAFT_BANNER = (
    "DRAFT (pre-publication marketing handoff) — values from the locked-config production run "
    "m2026r1 / cfg-20260611-production-lock (2026-06-13). Numbers are final/locked; the DRAFT "
    "mark reflects pending public layout and release, not provisional data."
)
# Header sub-label for this package. The shared PROVISIONAL_LABEL ("Pending Review — Subject to
# Change") contradicts the locked-final numbers, so this package uses a pre-publication label
# consistent with DRAFT_BANNER (numbers final; layout/publication pending).
PREPUB_LABEL = "Pre-publication marketing draft — numbers final/locked 2026-06-13; public layout pending"
SOURCE_CAPTION = "Source: ND State Data Center, locked-config baseline projection (2026-06-13)."
# Figure subtitle for chart/pyramid PNGs (was the now-stale "refresh after final production rerun").
FIGURE_SUBTITLE = "Pre-publication draft — ND State Data Center (2026 locked projection)"

# Locked-run provenance (see docs/.../final-run-metadata.md).
# Corrected full-horizon run: ADR-068 (2026-06-15) + 2026-06-16 survival-horizon amendment.
RUN_METHOD = "m2026r1"
RUN_CONFIG = "cfg-20260611-production-lock"  # functional config unchanged; ADR-068 deltas are comment-only + the survival build


def _git_short_commit() -> str:
    """Current short commit, resolved at runtime so workbook provenance never goes stale."""
    import subprocess

    try:
        return (
            subprocess.check_output(
                ["git", "rev-parse", "--short", "HEAD"],
                cwd=PROJECT_ROOT,
                stderr=subprocess.DEVNULL,
            )
            .decode()
            .strip()
            or "unknown"
        )
    except Exception:
        return "unknown"


RUN_COMMIT = _git_short_commit()
RUN_CONFIG_SHA16 = "a6e0bfbc2d70be85"  # sha256(16) of projection_config.yaml (2026-06-16; comment-only delta vs locked cca42fb42be76680)
RUN_DATE = "2026-06-16"


# ---------------------------------------------------------------------------
# Data loading
# ---------------------------------------------------------------------------


def _load_state(scenario: str) -> pd.DataFrame:
    """Load full single-year-age statewide parquet for a scenario."""
    path = (
        PROJECT_ROOT
        / "data"
        / "projections"
        / scenario
        / "state"
        / f"nd_state_38_projection_2025_2055_{scenario}.parquet"
    )
    df = pd.read_parquet(path)
    df["scenario"] = scenario
    return df


def _load_counties(scenario: str) -> pd.DataFrame:
    """Load all 53 county parquet files for a scenario; tag county_fips."""
    county_dir = PROJECT_ROOT / "data" / "projections" / scenario / "county"
    parts = []
    for pf in sorted(county_dir.glob("nd_county_*.parquet")):
        fips = pf.stem.split("_")[2]
        df = pd.read_parquet(pf)
        df["county_fips"] = fips
        parts.append(df)
    out = pd.concat(parts, ignore_index=True)
    out["scenario"] = scenario
    return out


def load_all() -> dict[str, dict[str, pd.DataFrame]]:
    """Return {scenario: {'state': df, 'county': df}} for public scenarios."""
    bundle: dict[str, dict[str, pd.DataFrame]] = {}
    for scen in PUBLIC_SCENARIOS:
        print(f"  loading {scen}...")
        bundle[scen] = {
            "state": _load_state(scen),
            "county": _load_counties(scen),
        }
    return bundle


# ---------------------------------------------------------------------------
# Aggregation to broad public age groups
# ---------------------------------------------------------------------------


def _broad_age_aggregate(df: pd.DataFrame, group_keys: list[str]) -> pd.DataFrame:
    """Aggregate to broad public age groups + sex totals.

    Returns one row per group_keys combination with columns:
        total_population, population_under_18, population_working_age_18_64,
        population_65_plus, population_85_plus, male_population,
        female_population, sex_ratio, youth_dependency_ratio,
        elderly_dependency_ratio, total_dependency_ratio.
    """
    base = df.groupby([*group_keys, "age", "sex"], observed=True)["population"].sum()
    base = base.reset_index()

    def _band(age: int) -> str:
        if age < 18:
            return "under_18"
        if age < 65:
            return "working_18_64"
        return "elderly_65_plus"

    base["band"] = base["age"].apply(_band)

    pivot = base.pivot_table(
        index=group_keys,
        columns="band",
        values="population",
        aggfunc="sum",
        fill_value=0.0,
    )
    for col in ("under_18", "working_18_64", "elderly_65_plus"):
        if col not in pivot.columns:
            pivot[col] = 0.0

    pivot_85 = (
        base[base["age"] >= 85]
        .groupby(group_keys, observed=True)["population"]
        .sum()
        .rename("pop_85_plus")
    )
    pivot = pivot.join(pivot_85, how="left").fillna(0.0)

    sex_pivot = base.pivot_table(
        index=group_keys,
        columns="sex",
        values="population",
        aggfunc="sum",
        fill_value=0.0,
    )
    for s in ("Male", "Female"):
        if s not in sex_pivot.columns:
            sex_pivot[s] = 0.0

    result = pivot.join(sex_pivot, how="left").fillna(0.0)
    result = result.reset_index()

    result["total_population"] = (
        result["under_18"] + result["working_18_64"] + result["elderly_65_plus"]
    )

    result["sex_ratio"] = (
        result["Male"] / result["Female"].where(result["Female"] > 0, other=pd.NA) * 100
    )
    result["sex_ratio"] = result["sex_ratio"].astype(float).round(1)

    working = result["working_18_64"].where(result["working_18_64"] > 0, other=pd.NA)
    result["youth_dependency_ratio"] = (result["under_18"] / working).astype(float).round(2)
    result["elderly_dependency_ratio"] = (
        (result["elderly_65_plus"] / working).astype(float).round(2)
    )
    result["total_dependency_ratio"] = (
        ((result["under_18"] + result["elderly_65_plus"]) / working).astype(float).round(2)
    )

    result = result.rename(
        columns={
            "under_18": "population_under_18",
            "working_18_64": "population_working_age_18_64",
            "elderly_65_plus": "population_65_plus",
            "pop_85_plus": "population_85_plus",
            "Male": "male_population",
            "Female": "female_population",
        }
    )
    return result


def build_tidy_public(bundle: dict[str, dict[str, pd.DataFrame]]) -> pd.DataFrame:
    """Build the tidy state+region+county dataset for the public baseline.

    Mirrors the schema in public-download-spec.md.
    """
    county_names = load_county_names()
    rows: list[pd.DataFrame] = []

    for scen in PUBLIC_SCENARIOS:
        state_df = bundle[scen]["state"]
        county_df = bundle[scen]["county"]

        # State
        s_agg = _broad_age_aggregate(state_df, ["scenario", "year"])
        s_agg["geography_level"] = "state"
        s_agg["geography_fips"] = "38"
        s_agg["geography_name"] = "North Dakota"
        s_agg["region_id"] = ""
        s_agg["region_name"] = ""
        rows.append(s_agg)

        # County
        c_agg = _broad_age_aggregate(county_df, ["scenario", "county_fips", "year"])
        c_agg["geography_level"] = "county"
        c_agg["geography_fips"] = c_agg["county_fips"]
        c_agg["geography_name"] = c_agg["county_fips"].map(
            lambda f: f"{county_names.get(f, f)} County"
        )
        c_agg["region_id"] = c_agg["county_fips"].map(lambda f: f"R{COUNTY_TO_REGION[f]}")
        c_agg["region_name"] = c_agg["county_fips"].map(lambda f: REGION_NAMES[COUNTY_TO_REGION[f]])

        # Region = sum of county broad aggregates by scenario, region, year
        c_with_region = c_agg.copy()
        region_agg_cols = [
            "total_population",
            "population_under_18",
            "population_working_age_18_64",
            "population_65_plus",
            "population_85_plus",
            "male_population",
            "female_population",
        ]
        r_agg = (
            c_with_region.groupby(["scenario", "region_id", "region_name", "year"], observed=True)[
                region_agg_cols
            ]
            .sum()
            .reset_index()
        )
        r_agg["geography_level"] = "region"
        r_agg["geography_fips"] = r_agg["region_id"]
        r_agg["geography_name"] = r_agg["region_name"]
        # Recompute ratios at region level
        r_agg["sex_ratio"] = (
            (
                r_agg["male_population"]
                / r_agg["female_population"].where(r_agg["female_population"] > 0, pd.NA)
                * 100
            )
            .astype(float)
            .round(1)
        )
        working = r_agg["population_working_age_18_64"].where(
            r_agg["population_working_age_18_64"] > 0, pd.NA
        )
        r_agg["youth_dependency_ratio"] = (
            (r_agg["population_under_18"] / working).astype(float).round(2)
        )
        r_agg["elderly_dependency_ratio"] = (
            (r_agg["population_65_plus"] / working).astype(float).round(2)
        )
        r_agg["total_dependency_ratio"] = (
            ((r_agg["population_under_18"] + r_agg["population_65_plus"]) / working)
            .astype(float)
            .round(2)
        )
        rows.append(r_agg)
        rows.append(c_agg.drop(columns="county_fips"))

    out = pd.concat(rows, ignore_index=True, sort=False)

    column_order = [
        "scenario",
        "geography_level",
        "geography_fips",
        "geography_name",
        "region_id",
        "region_name",
        "year",
        "total_population",
        "population_under_18",
        "population_working_age_18_64",
        "population_65_plus",
        "population_85_plus",
        "male_population",
        "female_population",
        "sex_ratio",
        "youth_dependency_ratio",
        "elderly_dependency_ratio",
        "total_dependency_ratio",
    ]
    out = out[column_order]
    out = out.sort_values(["scenario", "geography_level", "geography_fips", "year"]).reset_index(
        drop=True
    )
    return out


# ---------------------------------------------------------------------------
# Workbook helpers
# ---------------------------------------------------------------------------


def _write_header_block(ws, title: str, *, scenario_line: str | None = None) -> int:
    """Write a 4-row header block. Returns next available row."""
    ws.cell(row=1, column=1, value=title).font = HEADER_FONT
    if scenario_line:
        ws.cell(row=2, column=1, value=scenario_line).font = SUBTITLE_FONT
    ws.cell(row=3, column=1, value=DRAFT_BANNER).font = DRAFT_FONT
    ws.cell(row=4, column=1, value=PREPUB_LABEL).font = SUBTITLE_FONT
    return 6


def _write_dataframe(
    ws,
    df: pd.DataFrame,
    start_row: int,
    *,
    numeric_cols: set[str] | None = None,
    ratio_cols: set[str] | None = None,
    column_widths: dict[str, int] | None = None,
) -> int:
    """Write a DataFrame to ws starting at start_row. Returns next free row."""
    numeric_cols = numeric_cols or set()
    ratio_cols = ratio_cols or set()

    for ci, col in enumerate(df.columns, 1):
        cell = ws.cell(row=start_row, column=ci, value=col)
        cell.font = COL_HEADER_FONT
        cell.fill = COL_HEADER_FILL
        cell.alignment = COL_HEADER_ALIGN
    ws.freeze_panes = ws.cell(row=start_row + 1, column=1)

    for ri, row in enumerate(df.itertuples(index=False, name=None), start_row + 1):
        for ci, (col, value) in enumerate(zip(df.columns, row, strict=False), 1):
            if pd.isna(value):
                cell = ws.cell(row=ri, column=ci, value=None)
            elif col in numeric_cols:
                cell = ws.cell(row=ri, column=ci, value=round(float(value)))
                cell.number_format = NUM_FMT
            elif col in ratio_cols:
                cell = ws.cell(row=ri, column=ci, value=float(value))
                cell.number_format = RATIO_FMT
            else:
                cell = ws.cell(row=ri, column=ci, value=value)
            cell.font = NORMAL_FONT

    if column_widths:
        for col_name, width in column_widths.items():
            if col_name in df.columns:
                ci = list(df.columns).index(col_name) + 1
                ws.column_dimensions[get_column_letter(ci)].width = width

    return start_row + 1 + len(df)


# ---------------------------------------------------------------------------
# Workbook construction
# ---------------------------------------------------------------------------


def _build_readme_sheet(wb: Workbook) -> None:
    ws = wb.create_sheet("README")
    row = _write_header_block(
        ws,
        "North Dakota Population Projections 2025–2055 — Public Workbook (DRAFT)",
    )

    lines = [
        ("About this file", SECTION_FONT),
        (
            "Draft consolidated public workbook for the 2026 ND population "
            "projection release. Use for marketing layout and chart design.",
            NORMAL_FONT,
        ),
        ("", NORMAL_FONT),
        ("Run provenance", SECTION_FONT),
        (
            f"Method / config: {RUN_METHOD} / {RUN_CONFIG} "
            f"(alias county_champion). Run date: {RUN_DATE}.",
            NORMAL_FONT,
        ),
        (
            f"Git commit at run: {RUN_COMMIT}. "
            f"projection_config.yaml sha256 (16): {RUN_CONFIG_SHA16}.",
            NORMAL_FONT,
        ),
        (
            "Locked per ADR-061 (Accepted as modified), ADR-065 (CBO-adjusted "
            "baseline), ADR-066 (Vintage 2025 base), ADR-067 (Williams removed "
            "from college smoothing), ADR-068 (CBO migration numerator + open-ended "
            "90+ survival corrections, incl. the 2026-06-16 survival-horizon "
            "amendment). CF-001 disposed 2026-06-11.",
            NORMAL_FONT,
        ),
        ("", NORMAL_FONT),
        ("Status", SECTION_FONT),
        (
            "Numbers are final as of the corrected full-horizon production run "
            "(ADR-068, 2026-06-15 + 2026-06-16 survival-horizon amendment), which "
            "superseded the 2026-06-13 locked run. These files are a pre-publication "
            "marketing handoff: the DRAFT mark reflects pending public layout and "
            "release, not provisional data. Rerun only if the production outputs change.",
            NORMAL_FONT,
        ),
        ("", NORMAL_FONT),
        ("Coverage", SECTION_FONT),
        (
            "Geographies: 1 state + 8 economic planning regions (R1–R8) + "
            "53 counties. City/place projections are not part of the public "
            "release.",
            NORMAL_FONT,
        ),
        (
            "Years: 2025–2055, annual. Key years used in PDF: "
            "2025, 2030, 2035, 2040, 2045, 2050, 2055.",
            NORMAL_FONT,
        ),
        (
            "Scenario: Baseline (CBO-Adjusted). The former unadjusted "
            "trend-continuation and high-growth paths are internal sensitivity "
            "runs, not part of the public download.",
            NORMAL_FONT,
        ),
        ("", NORMAL_FONT),
        ("How to read these projections", SECTION_FONT),
        (CONDITIONAL_CAVEAT, NORMAL_FONT),
        (
            "Baseline carries the public narrative. It incorporates CBO-adjusted "
            "immigration and fertility assumptions and should be described as a "
            "projection, not a guaranteed outcome.",
            NORMAL_FONT,
        ),
        (
            "Population basis: totals include group-quarters (GQ) residents — "
            "dormitories, military barracks, nursing facilities — held constant at "
            "2025 levels. The model projects the household population and re-adds the "
            "constant GQ each year (ADR-055). Consequently any components of change "
            "(births, deaths, net migration) are HOUSEHOLD-BASIS and exclude GQ "
            "turnover; this workbook publishes population totals only, not components.",
            NORMAL_FONT,
        ),
        ("", NORMAL_FONT),
        ("Sheet guide", SECTION_FONT),
        ("README — this page.", NORMAL_FONT),
        (
            "State Key Years — statewide key-year baseline totals.",
            NORMAL_FONT,
        ),
        (
            "State Annual / Region Annual / County Annual — tidy long-format "
            "tables, one row per geography–year for the baseline scenario.",
            NORMAL_FONT,
        ),
        (
            "County Key Years — compact county table at the key years (baseline-led).",
            NORMAL_FONT,
        ),
        (
            "State Age Groups — under 18, 18–64, 65+, and 85+ totals by year.",
            NORMAL_FONT,
        ),
        (
            "State Age-Sex Detail — statewide population by 5-year age group "
            "(0–4 … 80–84, 85+) and sex at the key years 2025–2055.",
            NORMAL_FONT,
        ),
        (
            "Chart: … sheets — chart-ready cuts for the storyboard "
            "exhibits (statewide line, regional bars, county top/bottom, "
            "age trend, pyramids).",
            NORMAL_FONT,
        ),
        (
            "Data Dictionary — column definitions and scenario keys.",
            NORMAL_FONT,
        ),
        ("", NORMAL_FONT),
        ("Source", SECTION_FONT),
        (SOURCE_CAPTION, NORMAL_FONT),
        (
            "Produced by the ND State Data Center via "
            "scripts/exports/build_public_draft_package.py.",
            NORMAL_FONT,
        ),
    ]
    for text, font in lines:
        ws.cell(row=row, column=1, value=text).font = font
        row += 1

    ws.column_dimensions["A"].width = 100


def _build_state_key_years(wb: Workbook, tidy: pd.DataFrame) -> None:
    ws = wb.create_sheet("State Key Years")
    row = _write_header_block(ws, f"Statewide Baseline Key Years, {BASE_YEAR}–{FINAL_YEAR}")

    state = tidy[(tidy["geography_level"] == "state") & (tidy["year"].isin(KEY_YEARS))]
    pivot = state.pivot_table(
        index="scenario",
        columns="year",
        values="total_population",
        aggfunc="sum",
    ).reindex(PUBLIC_SCENARIOS)
    pivot.columns = [str(c) for c in pivot.columns]
    pivot.index = [SCENARIOS[s] for s in pivot.index]
    pivot[f"Change {BASE_YEAR}–{FINAL_YEAR}"] = pivot[str(FINAL_YEAR)] - pivot[str(BASE_YEAR)]
    pivot = pivot.reset_index().rename(columns={"index": "Scenario"})
    pivot = pivot.rename(columns={pivot.columns[0]: "Scenario"})

    numeric = {c for c in pivot.columns if c != "Scenario"}
    _write_dataframe(
        ws,
        pivot,
        row,
        numeric_cols=numeric,
        column_widths={"Scenario": 38},
    )

    for ci in range(2, len(pivot.columns) + 1):
        ws.column_dimensions[get_column_letter(ci)].width = 14


def _build_state_annual(wb: Workbook, tidy: pd.DataFrame) -> None:
    ws = wb.create_sheet("State Annual")
    row = _write_header_block(ws, "State Annual Projections (DRAFT)")
    df = tidy[tidy["geography_level"] == "state"].copy()
    df = df.drop(columns=["geography_level", "region_id", "region_name"])
    numeric = {
        "total_population",
        "population_under_18",
        "population_working_age_18_64",
        "population_65_plus",
        "population_85_plus",
        "male_population",
        "female_population",
    }
    ratios = {
        "sex_ratio",
        "youth_dependency_ratio",
        "elderly_dependency_ratio",
        "total_dependency_ratio",
    }
    _write_dataframe(
        ws,
        df,
        row,
        numeric_cols=numeric,
        ratio_cols=ratios,
        column_widths={"scenario": 22, "geography_name": 18},
    )


def _build_region_annual(wb: Workbook, tidy: pd.DataFrame) -> None:
    ws = wb.create_sheet("Region Annual")
    row = _write_header_block(ws, "Region Annual Projections (DRAFT)")
    df = tidy[tidy["geography_level"] == "region"].copy()
    df = df.drop(columns=["geography_level"])
    numeric = {
        "total_population",
        "population_under_18",
        "population_working_age_18_64",
        "population_65_plus",
        "population_85_plus",
        "male_population",
        "female_population",
    }
    ratios = {
        "sex_ratio",
        "youth_dependency_ratio",
        "elderly_dependency_ratio",
        "total_dependency_ratio",
    }
    _write_dataframe(
        ws,
        df,
        row,
        numeric_cols=numeric,
        ratio_cols=ratios,
        column_widths={"scenario": 22, "geography_name": 16, "region_name": 16},
    )


def _build_county_annual(wb: Workbook, tidy: pd.DataFrame) -> None:
    ws = wb.create_sheet("County Annual")
    row = _write_header_block(ws, "County Annual Projections (DRAFT)")
    df = tidy[tidy["geography_level"] == "county"].copy()
    df = df.drop(columns=["geography_level"])
    numeric = {
        "total_population",
        "population_under_18",
        "population_working_age_18_64",
        "population_65_plus",
        "population_85_plus",
        "male_population",
        "female_population",
    }
    ratios = {
        "sex_ratio",
        "youth_dependency_ratio",
        "elderly_dependency_ratio",
        "total_dependency_ratio",
    }
    _write_dataframe(
        ws,
        df,
        row,
        numeric_cols=numeric,
        ratio_cols=ratios,
        column_widths={"scenario": 22, "geography_name": 22, "region_name": 16},
    )


def _build_county_key_years(wb: Workbook, tidy: pd.DataFrame) -> None:
    ws = wb.create_sheet("County Key Years")
    row = _write_header_block(
        ws,
        "County Key-Year Baseline Totals (DRAFT)",
        scenario_line="Scenario: Baseline (CBO-Adjusted)",
    )
    df = tidy[
        (tidy["geography_level"] == "county")
        & (tidy["scenario"] == "baseline")
        & (tidy["year"].isin(KEY_YEARS))
    ].copy()
    pivot = df.pivot_table(
        index=["geography_fips", "geography_name", "region_name"],
        columns="year",
        values="total_population",
        aggfunc="sum",
    ).reset_index()
    pivot.columns = [str(c) for c in pivot.columns]
    pivot[f"Change {BASE_YEAR}–{FINAL_YEAR}"] = pivot[str(FINAL_YEAR)] - pivot[str(BASE_YEAR)]
    pivot = pivot.sort_values("geography_name")
    numeric = {
        c for c in pivot.columns if c not in {"geography_fips", "geography_name", "region_name"}
    }
    _write_dataframe(
        ws,
        pivot,
        row,
        numeric_cols=numeric,
        column_widths={
            "geography_fips": 12,
            "geography_name": 22,
            "region_name": 16,
        },
    )


def _build_state_age_groups(wb: Workbook, tidy: pd.DataFrame) -> None:
    ws = wb.create_sheet("State Age Groups")
    row = _write_header_block(ws, "Baseline State Age Group Totals (DRAFT)")
    df = tidy[tidy["geography_level"] == "state"].copy()
    keep = [
        "scenario",
        "year",
        "population_under_18",
        "population_working_age_18_64",
        "population_65_plus",
        "population_85_plus",
    ]
    df = df[keep].rename(
        columns={
            "population_under_18": "Under 18",
            "population_working_age_18_64": "18-64",
            "population_65_plus": "65+",
            "population_85_plus": "85+",
        }
    )
    numeric = {"Under 18", "18-64", "65+", "85+"}
    _write_dataframe(
        ws,
        df,
        row,
        numeric_cols=numeric,
        column_widths={"scenario": 22},
    )


def _build_state_age_sex_detail(
    wb: Workbook, bundle: dict[str, dict[str, pd.DataFrame]]
) -> None:
    """State population by 5-year age group and sex at the key years.

    One row per (age group, sex); one column per key year. Built from the
    single-year-age state parquet so the detail is exact, then binned into the
    standard 5-year groups (0-4 … 80-84, 85+).
    """
    ws = wb.create_sheet("State Age-Sex Detail")
    row = _write_header_block(
        ws,
        "Baseline State Population by 5-Year Age Group and Sex (DRAFT)",
        scenario_line="Scenario: Baseline (CBO-Adjusted); key years 2025–2055",
    )

    state = bundle["baseline"]["state"]
    df = state[state["year"].isin(KEY_YEARS)].copy()
    df["age_group_start"] = ((df["age"] // 5) * 5).clip(upper=85)

    def _label(start: int) -> str:
        return "85+" if start >= 85 else f"{int(start)}-{int(start) + 4}"

    df["Age Group"] = df["age_group_start"].apply(_label)

    grouped = (
        df.groupby(["age_group_start", "Age Group", "sex", "year"], observed=True)["population"]
        .sum()
        .reset_index()
    )
    pivot = grouped.pivot_table(
        index=["age_group_start", "Age Group", "sex"],
        columns="year",
        values="population",
        aggfunc="sum",
        fill_value=0.0,
    ).reset_index()
    pivot = pivot.sort_values(["age_group_start", "sex"]).drop(columns="age_group_start")
    pivot = pivot.rename(columns={"sex": "Sex"})
    pivot.columns = [str(c) if isinstance(c, int) else c for c in pivot.columns]
    year_cols = [str(y) for y in KEY_YEARS]
    pivot = pivot[["Age Group", "Sex", *year_cols]]

    _write_dataframe(
        ws,
        pivot,
        row,
        numeric_cols=set(year_cols),
        column_widths={"Age Group": 14, "Sex": 10},
    )
    for ci in range(3, len(pivot.columns) + 1):
        ws.column_dimensions[get_column_letter(ci)].width = 12


def _build_chart_state_line(wb: Workbook, tidy: pd.DataFrame) -> None:
    ws = wb.create_sheet("Chart - State Line")
    row = _write_header_block(
        ws,
        "Statewide Baseline Population (chart-ready)",
        scenario_line="Storyboard page 5: 2025–2055 baseline line chart",
    )
    df = tidy[tidy["geography_level"] == "state"][["scenario", "year", "total_population"]].copy()
    df["scenario_label"] = df["scenario"].map(SCENARIOS)
    df = df[["scenario", "scenario_label", "year", "total_population"]]
    _write_dataframe(
        ws,
        df,
        row,
        numeric_cols={"total_population"},
        column_widths={"scenario": 22, "scenario_label": 38},
    )


def _build_chart_region_bars(wb: Workbook, tidy: pd.DataFrame) -> None:
    ws = wb.create_sheet("Chart - Region Bars")
    row = _write_header_block(
        ws,
        "Baseline Regional Change 2025–2055 (chart-ready)",
        scenario_line="Storyboard page 7: regional change bar chart",
    )
    df = tidy[
        (tidy["geography_level"] == "region")
        & (tidy["scenario"] == "baseline")
        & (tidy["year"].isin([BASE_YEAR, FINAL_YEAR]))
    ].copy()
    pivot = df.pivot_table(
        index=["region_id", "region_name"],
        columns="year",
        values="total_population",
        aggfunc="sum",
    ).reset_index()
    pivot.columns = [str(c) if isinstance(c, int) else c for c in pivot.columns]
    pivot["change"] = pivot[str(FINAL_YEAR)] - pivot[str(BASE_YEAR)]
    pivot = pivot.sort_values("change", ascending=False)
    _write_dataframe(
        ws,
        pivot,
        row,
        numeric_cols={str(BASE_YEAR), str(FINAL_YEAR), "change"},
        column_widths={"region_id": 10, "region_name": 16},
    )


def _build_chart_county_top_bottom(wb: Workbook, tidy: pd.DataFrame) -> None:
    ws = wb.create_sheet("Chart - County Top-Bottom")
    row = _write_header_block(
        ws,
        "Top 6 Growth and Bottom 6 Decline Counties, Baseline (chart-ready)",
        scenario_line="Storyboard page 8: county growth/decline bar chart",
    )
    df = tidy[
        (tidy["geography_level"] == "county")
        & (tidy["scenario"] == "baseline")
        & (tidy["year"].isin([BASE_YEAR, FINAL_YEAR]))
    ].copy()
    pivot = df.pivot_table(
        index=["geography_fips", "geography_name", "region_name"],
        columns="year",
        values="total_population",
        aggfunc="sum",
    ).reset_index()
    pivot.columns = [str(c) if isinstance(c, int) else c for c in pivot.columns]
    pivot["change"] = pivot[str(FINAL_YEAR)] - pivot[str(BASE_YEAR)]
    pivot = pivot.sort_values("change", ascending=False)
    top = pivot.head(6).assign(group="Top growth")
    bottom = pivot.tail(6).assign(group="Largest decline")
    out = pd.concat([top, bottom], ignore_index=True)
    out = out[
        [
            "group",
            "geography_fips",
            "geography_name",
            "region_name",
            str(BASE_YEAR),
            str(FINAL_YEAR),
            "change",
        ]
    ]
    _write_dataframe(
        ws,
        out,
        row,
        numeric_cols={str(BASE_YEAR), str(FINAL_YEAR), "change"},
        column_widths={
            "group": 18,
            "geography_fips": 10,
            "geography_name": 22,
            "region_name": 16,
        },
    )


def _build_chart_age_trend(wb: Workbook, tidy: pd.DataFrame) -> None:
    ws = wb.create_sheet("Chart - Age Trend")
    row = _write_header_block(
        ws,
        "Baseline Age Group Trends 2025–2055 (chart-ready)",
        scenario_line="Storyboard page 9: age group trend line chart",
    )
    df = tidy[(tidy["geography_level"] == "state") & (tidy["scenario"] == "baseline")].copy()
    out = df[
        [
            "year",
            "population_under_18",
            "population_working_age_18_64",
            "population_65_plus",
            "population_85_plus",
        ]
    ].rename(
        columns={
            "population_under_18": "Under 18",
            "population_working_age_18_64": "18-64",
            "population_65_plus": "65+",
            "population_85_plus": "85+",
        }
    )
    _write_dataframe(
        ws,
        out,
        row,
        numeric_cols={"Under 18", "18-64", "65+", "85+"},
    )


def _build_chart_pyramid(
    wb: Workbook,
    state_df: pd.DataFrame,
    year: int,
    scenario_key: str,
    sheet_name: str,
) -> None:
    ws = wb.create_sheet(sheet_name)
    row = _write_header_block(
        ws,
        f"Pyramid Data — State, {year}, {SCENARIOS[scenario_key]} (chart-ready)",
        scenario_line="Storyboard page 10: pyramid chart data",
    )
    one = state_df[state_df["year"] == year].copy()
    one["age_group_start"] = (one["age"] // 5) * 5

    def _label(start: int) -> str:
        if start >= 85:
            return "85+"
        return f"{start}-{start + 4}"

    one["age_group"] = one["age_group_start"].apply(_label)

    grouped = (
        one.groupby(["age_group_start", "age_group", "sex"], observed=True)["population"]
        .sum()
        .reset_index()
    )
    pivot = grouped.pivot_table(
        index=["age_group_start", "age_group"],
        columns="sex",
        values="population",
        aggfunc="sum",
        fill_value=0.0,
    ).reset_index()
    pivot = pivot.sort_values("age_group_start").drop(columns="age_group_start")
    for col in ("Male", "Female"):
        if col not in pivot.columns:
            pivot[col] = 0.0
    pivot = pivot[["age_group", "Male", "Female"]]
    _write_dataframe(
        ws,
        pivot,
        row,
        numeric_cols={"Male", "Female"},
        column_widths={"age_group": 14},
    )


def _build_data_dictionary(wb: Workbook) -> None:
    ws = wb.create_sheet("Data Dictionary")
    row = _write_header_block(ws, "Data Dictionary (DRAFT)")

    rows = [
        ("Column", "Type", "Description"),
        ("scenario", "string", "Scenario key: baseline"),
        ("geography_level", "string", "state, region, or county"),
        ("geography_fips", "string", "State or county FIPS; region code for region rows"),
        ("geography_name", "string", "Display name"),
        ("region_id", "string", "Region ID R1-R8 for region/county rows; blank for state"),
        ("region_name", "string", "Region display name for region/county rows; blank for state"),
        ("year", "integer", "Projection year, 2025-2055"),
        ("total_population", "number", "Projected total population"),
        ("population_under_18", "number", "Projected population under age 18"),
        ("population_working_age_18_64", "number", "Projected population age 18-64"),
        ("population_65_plus", "number", "Projected population age 65 and older"),
        ("population_85_plus", "number", "Projected population age 85 and older"),
        ("male_population", "number", "Projected male population"),
        ("female_population", "number", "Projected female population"),
        ("sex_ratio", "number", "Males per 100 females"),
        (
            "youth_dependency_ratio",
            "number",
            "population_under_18 / population_working_age_18_64",
        ),
        (
            "elderly_dependency_ratio",
            "number",
            "population_65_plus / population_working_age_18_64",
        ),
        (
            "total_dependency_ratio",
            "number",
            "(population_under_18 + population_65_plus) / population_working_age_18_64",
        ),
    ]
    for ri, row_vals in enumerate(rows):
        for ci, val in enumerate(row_vals, 1):
            cell = ws.cell(row=row + ri, column=ci, value=val)
            if ri == 0:
                cell.font = COL_HEADER_FONT
                cell.fill = COL_HEADER_FILL
                cell.alignment = COL_HEADER_ALIGN
            else:
                cell.font = NORMAL_FONT
    row += len(rows) + 2

    ws.cell(row=row, column=1, value="Scenario keys").font = SECTION_FONT
    row += 1
    for key in PUBLIC_SCENARIOS:
        label = SCENARIOS[key]
        ws.cell(row=row, column=1, value=key).font = NORMAL_FONT
        ws.cell(row=row, column=2, value=label).font = NORMAL_FONT
        row += 1
    row += 1

    ws.cell(row=row, column=1, value="Region keys").font = SECTION_FONT
    row += 1
    for num in sorted(REGION_NAMES):
        ws.cell(row=row, column=1, value=f"R{num}").font = NORMAL_FONT
        ws.cell(row=row, column=2, value=REGION_NAMES[num]).font = NORMAL_FONT
        county_count = len(REGION_COUNTIES[num])
        ws.cell(row=row, column=3, value=f"{county_count} counties").font = NORMAL_FONT
        row += 1

    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 80


def build_workbook(
    tidy: pd.DataFrame,
    bundle: dict[str, dict[str, pd.DataFrame]],
    output_path: Path,
) -> None:
    """Assemble the consolidated draft public workbook."""
    wb = Workbook()
    default = wb.active
    if default is not None:
        wb.remove(default)

    _build_readme_sheet(wb)
    _build_state_key_years(wb, tidy)
    _build_state_annual(wb, tidy)
    _build_region_annual(wb, tidy)
    _build_county_annual(wb, tidy)
    _build_county_key_years(wb, tidy)
    _build_state_age_groups(wb, tidy)
    _build_state_age_sex_detail(wb, bundle)
    _build_chart_state_line(wb, tidy)
    _build_chart_region_bars(wb, tidy)
    _build_chart_county_top_bottom(wb, tidy)
    _build_chart_age_trend(wb, tidy)
    _build_chart_pyramid(
        wb,
        bundle["baseline"]["state"],
        BASE_YEAR,
        "baseline",
        "Chart - Pyramid 2025",
    )
    _build_chart_pyramid(
        wb,
        bundle["baseline"]["state"],
        FINAL_YEAR,
        "baseline",
        "Chart - Pyramid 2055 Baseline",
    )
    _build_data_dictionary(wb)

    wb.save(output_path)
    print(f"  workbook → {output_path.relative_to(PROJECT_ROOT)}")


def write_public_csv(tidy: pd.DataFrame, output_path: Path) -> None:
    """Write the consolidated long-format public CSV download."""

    tidy.to_csv(output_path, index=False)
    print(f"  csv → {output_path.relative_to(PROJECT_ROOT)}")


def remove_stale_alternative_scenario_artifacts(output_dir: Path) -> None:
    """Remove draft artifacts from the former multi-scenario public package."""

    removed = []
    for filename in STALE_ALTERNATIVE_SCENARIO_ARTIFACTS:
        path = output_dir / filename
        if path.exists():
            path.unlink()
            removed.append(filename)

    if removed:
        print(f"  removed stale non-public scenario artifacts: {', '.join(removed)}")


# ---------------------------------------------------------------------------
# Pyramid PNG generation
# ---------------------------------------------------------------------------


def _make_pyramid(
    df: pd.DataFrame,
    year: int,
    output_path: Path,
    title: str,
) -> None:
    from cohort_projections.output.visualizations import plot_population_pyramid

    plot_population_pyramid(
        projection_df=df,
        year=year,
        output_path=output_path,
        by_race=False,
        age_group_size=5,
        title=f"{title}\n{FIGURE_SUBTITLE}",
        figsize=(9, 7),
        dpi=300,
    )


def generate_pyramids(
    bundle: dict[str, dict[str, pd.DataFrame]],
    output_dir: Path,
) -> None:
    """Generate the baseline reference pyramid PNGs."""
    state = bundle["baseline"]["state"]
    _make_pyramid(
        state,
        BASE_YEAR,
        output_dir / "pyramid_state_2025.png",
        f"North Dakota Population Pyramid {BASE_YEAR}",
    )
    _make_pyramid(
        state,
        FINAL_YEAR,
        output_dir / "pyramid_state_2055_baseline.png",
        f"North Dakota Population Pyramid {FINAL_YEAR} — Baseline",
    )
    for fips, label in [("38017", "Cass County"), ("38101", "Ward County")]:
        county_df = bundle["baseline"]["county"]
        sub = county_df[county_df["county_fips"] == fips]
        slug = label.lower().split()[0]
        _make_pyramid(
            sub,
            BASE_YEAR,
            output_dir / f"pyramid_{slug}_2025.png",
            f"{label} Population Pyramid {BASE_YEAR}",
        )
        _make_pyramid(
            sub,
            FINAL_YEAR,
            output_dir / f"pyramid_{slug}_2055_baseline.png",
            f"{label} Population Pyramid {FINAL_YEAR} — Baseline",
        )

    print(f"  pyramids → {output_dir.relative_to(PROJECT_ROOT)} (6 images)")


# ---------------------------------------------------------------------------
# Key reference chart PNGs
# ---------------------------------------------------------------------------


def _draft_figure_style(ax, title: str, *, ylabel: str | None = None) -> None:
    ax.set_title(
        f"{title}\n{FIGURE_SUBTITLE}",
        fontsize=13,
        fontweight="bold",
    )
    if ylabel:
        ax.set_ylabel(ylabel)
    ax.grid(True, axis="y", linestyle=":", alpha=0.5)


def generate_key_charts(tidy: pd.DataFrame, output_dir: Path) -> None:
    """Produce the 4 reference key-chart PNGs."""
    import matplotlib

    matplotlib.use("Agg")
    import matplotlib.pyplot as plt
    import matplotlib.ticker as mticker

    # 1. State baseline line
    fig, ax = plt.subplots(figsize=(10, 6), dpi=300)
    state = tidy[tidy["geography_level"] == "state"]
    sub = state[state["scenario"] == "baseline"].sort_values("year")
    ax.plot(
        sub["year"],
        sub["total_population"],
        label=SCENARIOS["baseline"],
        color="#1F3864",
        linewidth=2.5,
        marker="o",
        markersize=4,
    )
    ax.yaxis.set_major_formatter(mticker.FuncFormatter(lambda v, _p: f"{int(v):,}"))
    ax.set_xlabel("Year")
    ax.legend(loc="upper left", frameon=True)
    _draft_figure_style(
        ax,
        "North Dakota Statewide Baseline Population, 2025–2055",
        ylabel="Population",
    )
    fig.tight_layout()
    fig.savefig(output_dir / "chart_state_baseline_line.png", bbox_inches="tight")
    plt.close(fig)

    # 2. Regional baseline change bars
    fig, ax = plt.subplots(figsize=(10, 6), dpi=300)
    region = tidy[
        (tidy["geography_level"] == "region")
        & (tidy["scenario"] == "baseline")
        & (tidy["year"].isin([BASE_YEAR, FINAL_YEAR]))
    ]
    pivot = region.pivot_table(
        index=["region_id", "region_name"],
        columns="year",
        values="total_population",
        aggfunc="sum",
    ).reset_index()
    pivot["change"] = pivot[FINAL_YEAR] - pivot[BASE_YEAR]
    pivot = pivot.sort_values("change", ascending=True)
    bar_colors = ["#C62828" if c < 0 else "#1F3864" for c in pivot["change"]]
    ax.barh(pivot["region_name"], pivot["change"], color=bar_colors)
    ax.axvline(0, color="black", linewidth=0.7)
    ax.xaxis.set_major_formatter(mticker.FuncFormatter(lambda v, _p: f"{int(v):,}"))
    _draft_figure_style(
        ax,
        "Baseline Regional Change 2025–2055",
        ylabel="Region",
    )
    ax.set_xlabel("Change in population (residents)")
    fig.tight_layout()
    fig.savefig(output_dir / "chart_region_baseline_bars.png", bbox_inches="tight")
    plt.close(fig)

    # 3. County top growth / bottom decline
    fig, ax = plt.subplots(figsize=(10, 7), dpi=300)
    county = tidy[
        (tidy["geography_level"] == "county")
        & (tidy["scenario"] == "baseline")
        & (tidy["year"].isin([BASE_YEAR, FINAL_YEAR]))
    ]
    cpivot = county.pivot_table(
        index=["geography_fips", "geography_name"],
        columns="year",
        values="total_population",
        aggfunc="sum",
    ).reset_index()
    cpivot["change"] = cpivot[FINAL_YEAR] - cpivot[BASE_YEAR]
    cpivot = cpivot.sort_values("change", ascending=False)
    top6 = cpivot.head(6)
    bot6 = cpivot.tail(6).sort_values("change", ascending=True)
    combined = pd.concat([bot6, top6], ignore_index=True)
    bar_colors = ["#C62828" if c < 0 else "#1F3864" for c in combined["change"]]
    ax.barh(combined["geography_name"], combined["change"], color=bar_colors)
    ax.axvline(0, color="black", linewidth=0.7)
    ax.xaxis.set_major_formatter(mticker.FuncFormatter(lambda v, _p: f"{int(v):,}"))
    _draft_figure_style(
        ax,
        "Top 6 Growth and Bottom 6 Decline Counties, Baseline 2025–2055",
        ylabel="County",
    )
    ax.set_xlabel("Change in population (residents)")
    fig.tight_layout()
    fig.savefig(output_dir / "chart_county_top_bottom.png", bbox_inches="tight")
    plt.close(fig)

    # 4. Age group trend
    fig, ax = plt.subplots(figsize=(10, 6), dpi=300)
    age = tidy[(tidy["geography_level"] == "state") & (tidy["scenario"] == "baseline")].sort_values(
        "year"
    )
    series = [
        ("Under 18", age["population_under_18"], "#2E7D32"),
        ("18–64", age["population_working_age_18_64"], "#1F3864"),
        ("65+", age["population_65_plus"], "#E65100"),
        ("85+", age["population_85_plus"], "#6A1B9A"),
    ]
    for label, vals, color in series:
        ax.plot(age["year"], vals, label=label, color=color, linewidth=2, marker="o", markersize=3)
    ax.yaxis.set_major_formatter(mticker.FuncFormatter(lambda v, _p: f"{int(v):,}"))
    ax.set_xlabel("Year")
    ax.legend(loc="upper left", frameon=True)
    _draft_figure_style(
        ax,
        "Baseline Statewide Age Group Trends, 2025–2055",
        ylabel="Population",
    )
    fig.tight_layout()
    fig.savefig(output_dir / "chart_age_group_trend.png", bbox_inches="tight")
    plt.close(fig)

    print(f"  key charts → {output_dir.relative_to(PROJECT_ROOT)} (4 images)")


# ---------------------------------------------------------------------------
# Bundled reference PDF
# ---------------------------------------------------------------------------


def copy_reference_pdf() -> None:
    src = PROJECT_ROOT / "data" / "raw" / "ND Population Projections.pdf"
    dst = MARKETING_DIR / REFERENCE_PDF_NAME
    if not src.exists():
        print(f"  WARNING: 2024 SDC PDF not found at {src}")
        return
    shutil.copy2(src, dst)
    print(f"  reference PDF → {dst.relative_to(PROJECT_ROOT)}")


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------


def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument(
        "--skip-charts",
        action="store_true",
        help="Skip pyramid + key chart generation (workbook + PDF copy only).",
    )
    args = parser.parse_args()

    DRAFTS_DIR.mkdir(parents=True, exist_ok=True)
    print(f"Output directory: {DRAFTS_DIR.relative_to(PROJECT_ROOT)}")
    remove_stale_alternative_scenario_artifacts(DRAFTS_DIR)

    print("Loading projection parquets...")
    bundle = load_all()

    print("Building tidy public dataset...")
    tidy = build_tidy_public(bundle)
    print(f"  tidy rows: {len(tidy):,}")

    print("Building consolidated workbook...")
    build_workbook(tidy, bundle, DRAFTS_DIR / WORKBOOK_NAME)

    print("Writing consolidated CSV...")
    write_public_csv(tidy, DRAFTS_DIR / CSV_NAME)

    if not args.skip_charts:
        print("Generating pyramids...")
        generate_pyramids(bundle, DRAFTS_DIR)

        print("Generating key reference charts...")
        generate_key_charts(tidy, DRAFTS_DIR)

    print("Copying 2024 SDC reference PDF...")
    copy_reference_pdf()

    print("Done.")


if __name__ == "__main__":
    main()
