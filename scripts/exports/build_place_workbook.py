#!/usr/bin/env python3
"""
Build standalone place workbooks for PP-003 (IMP-14).

Creates one workbook per scenario:
`nd_projections_{scenario}_places_{datestamp}.xlsx`

Workbook structure follows S06 Section 7.2:
1. Table of Contents
2. 9 HIGH-tier sheets (age-group x sex at key years)
3. 9 MODERATE-tier sheets (broad age-group x sex at key years)
4. 1 LOWER-tier combined sheet (total population at key years)
5. HU Comparison (optional, if housing-unit projections available)
6. Methodology
"""

from __future__ import annotations

import argparse
import logging
import re
import sys
from datetime import UTC, datetime
from pathlib import Path
from typing import Any

import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

project_root = Path(__file__).resolve().parents[2]
sys.path.insert(0, str(project_root))
sys.path.insert(0, str(Path(__file__).parent))

from _methodology import (  # noqa: E402
    CONDITIONAL_CAVEAT,
    METHODOLOGY_LINES,
    ORGANIZATION_ATTRIBUTION,
    PLACE_METHODOLOGY_LINE,
    PROVISIONAL_LABEL,
    SCENARIOS,
)

from cohort_projections.utils import load_projection_config  # noqa: E402

TODAY = datetime.now(tz=UTC).date()
DATE_STAMP = TODAY.strftime("%Y%m%d")
DEFAULT_KEY_YEARS = [2025, 2030, 2035, 2040, 2045, 2050, 2055]
TIER_ORDER = ["HIGH", "MODERATE", "LOWER"]
HIGH_AGE_GROUPS = [
    "0-4",
    "5-9",
    "10-14",
    "15-19",
    "20-24",
    "25-29",
    "30-34",
    "35-39",
    "40-44",
    "45-49",
    "50-54",
    "55-59",
    "60-64",
    "65-69",
    "70-74",
    "75-79",
    "80-84",
    "85+",
]
MODERATE_AGE_GROUPS = ["0-17", "18-24", "25-44", "45-64", "65-84", "85+"]

HEADER_FONT = Font(name="Aptos", size=14, bold=True, color="1F3864")
SUBTITLE_FONT = Font(name="Aptos", size=11, italic=True, color="595959")
PROVISIONAL_FONT = Font(name="Aptos", size=11, bold=True, color="C00000")
SECTION_FONT = Font(name="Aptos", size=11, bold=True, color="1F3864")
COL_HEADER_FONT = Font(name="Aptos", size=10, bold=True, color="FFFFFF")
COL_HEADER_FILL = PatternFill("solid", fgColor="1F3864")
NORMAL_FONT = Font(name="Aptos", size=10)
LINK_FONT = Font(name="Aptos", size=10, color="0563C1", underline="single")
METHODOLOGY_FONT = Font(name="Aptos", size=10, color="595959")
CAVEAT_FONT = Font(name="Aptos", size=11, bold=True, color="C00000")
THIN_BORDER = Border(bottom=Side(style="thin", color="D9D9D9"))
NUM_FMT = "#,##0"
PCT_FMT = "0.0%"

logger = logging.getLogger(__name__)


def _resolve_path(path_value: str | Path, root: Path) -> Path:
    """Resolve config path values to absolute filesystem paths."""
    path = Path(path_value)
    return path if path.is_absolute() else root / path


def _resolve_scenarios(config: dict[str, Any], requested: list[str] | None) -> list[str]:
    """Resolve scenarios from CLI override or active config entries."""
    if requested:
        return requested

    config_scenarios = config.get("scenarios", {})
    active = [
        name
        for name, settings in config_scenarios.items()
        if isinstance(settings, dict) and settings.get("active", False)
    ]
    if active:
        return active

    fallback = config.get("pipeline", {}).get("projection", {}).get("scenarios", ["baseline"])
    if isinstance(fallback, list) and fallback:
        return [str(scenario) for scenario in fallback]
    return ["baseline"]


def _key_years(config: dict[str, Any]) -> list[int]:
    """Return configured place key years (fallback to defaults)."""
    configured = config.get("place_projections", {}).get("output", {}).get("key_years", DEFAULT_KEY_YEARS)
    key_years = [int(year) for year in configured]
    return key_years if key_years else list(DEFAULT_KEY_YEARS)


def _projection_root(config: dict[str, Any]) -> Path:
    """Return projection output root from config."""
    output_dir = config.get("pipeline", {}).get("projection", {}).get("output_dir", "data/projections")
    return _resolve_path(output_dir, project_root)


def _export_root(config: dict[str, Any]) -> Path:
    """Return export output root from config."""
    output_dir = config.get("pipeline", {}).get("export", {}).get("output_dir", "data/exports")
    return _resolve_path(output_dir, project_root)


def load_county_names() -> dict[str, str]:
    """Load county FIPS to display-name mapping from ND county reference CSV."""
    county_csv = project_root / "data" / "raw" / "geographic" / "nd_counties.csv"
    if not county_csv.exists():
        logger.warning("County reference file not found: %s", county_csv)
        return {}

    county_df = pd.read_csv(county_csv, dtype=str)
    nd = county_df[county_df["state_fips"] == "38"]
    names: dict[str, str] = {}
    for _, row in nd.iterrows():
        fips = str(row["county_fips"]).zfill(5)
        names[fips] = str(row["county_name"]).removesuffix(" County")
    return names


def _clean_sheet_title(name: str) -> str:
    """Remove invalid Excel worksheet characters and collapse whitespace."""
    cleaned = re.sub(r"[\\/*?:\[\]]", " ", str(name))
    cleaned = re.sub(r"\s+", " ", cleaned).strip()
    return cleaned or "Sheet"


def _unique_sheet_name(base: str, used_names: set[str]) -> str:
    """Create a unique, Excel-safe sheet name (<=31 chars)."""
    clean_base = _clean_sheet_title(base)
    candidate = clean_base[:31]
    if candidate not in used_names:
        used_names.add(candidate)
        return candidate

    suffix_index = 2
    while True:
        suffix = f" ({suffix_index})"
        truncated = clean_base[: max(1, 31 - len(suffix))]
        candidate = f"{truncated}{suffix}"
        if candidate not in used_names:
            used_names.add(candidate)
            return candidate
        suffix_index += 1


def _load_places_summary(scenario: str, config: dict[str, Any]) -> pd.DataFrame:
    """Load place summary rows for one scenario."""
    summary_path = _projection_root(config) / scenario / "place" / "places_summary.csv"
    if not summary_path.exists():
        raise FileNotFoundError(f"Missing places summary for {scenario}: {summary_path}")

    summary_df = pd.read_csv(summary_path, dtype={"place_fips": str, "county_fips": str})
    required_columns = {
        "place_fips",
        "name",
        "county_fips",
        "row_type",
        "confidence_tier",
        "base_population",
        "final_population",
        "growth_rate",
    }
    missing = required_columns - set(summary_df.columns)
    if missing:
        raise ValueError(f"places_summary missing required columns: {sorted(missing)}")

    place_rows = summary_df[summary_df["row_type"] == "place"].copy()
    place_rows["confidence_tier"] = place_rows["confidence_tier"].astype(str).str.upper()
    place_rows = place_rows[place_rows["confidence_tier"].isin(TIER_ORDER)].copy()
    if place_rows.empty:
        raise ValueError(f"No projected place rows found in {summary_path}")

    for numeric_col in ["base_population", "final_population", "growth_rate"]:
        place_rows[numeric_col] = pd.to_numeric(place_rows[numeric_col], errors="coerce").fillna(0.0)

    place_rows["name"] = place_rows["name"].astype(str)
    place_rows["place_fips"] = place_rows["place_fips"].astype(str).str.zfill(7)
    place_rows["county_fips"] = place_rows["county_fips"].astype(str).str.zfill(5)
    return place_rows.sort_values(["confidence_tier", "name", "place_fips"]).reset_index(drop=True)


def _find_place_projection_path(scenario: str, place_fips: str, config: dict[str, Any]) -> Path:
    """Locate one per-place projection parquet file."""
    place_dir = _projection_root(config) / scenario / "place"
    matches = sorted(place_dir.glob(f"nd_place_{place_fips}_projection_*_{scenario}.parquet"))
    if not matches:
        raise FileNotFoundError(f"Missing place projection parquet for {scenario}, place {place_fips}")
    return matches[0]


def _load_place_projection(scenario: str, place_fips: str, config: dict[str, Any]) -> pd.DataFrame:
    """Load one place projection parquet file."""
    parquet_path = _find_place_projection_path(scenario, place_fips, config)
    projection_df = pd.read_parquet(parquet_path)
    projection_df["year"] = pd.to_numeric(projection_df["year"], errors="coerce").fillna(-1).astype(int)
    if "population" in projection_df.columns:
        projection_df["population"] = pd.to_numeric(projection_df["population"], errors="coerce").fillna(0.0)
    return projection_df


def _totals_by_year(place_df: pd.DataFrame, key_years: list[int]) -> dict[int, float]:
    """Return total population by key year from tier-specific place output."""
    if "population" not in place_df.columns:
        return dict.fromkeys(key_years, 0.0)
    totals = place_df.groupby("year", as_index=True)["population"].sum()
    return {year: float(totals.get(year, 0.0)) for year in key_years}


def _write_header_block(
    ws: Any,
    title: str,
    scenario_label: str,
    subtitle: str | None = None,
) -> int:
    """Write standard workbook sheet header and return next row."""
    ws.cell(row=1, column=1, value=title).font = HEADER_FONT
    ws.cell(row=2, column=1, value=f"Scenario: {scenario_label}").font = SUBTITLE_FONT
    if subtitle:
        ws.cell(row=3, column=1, value=subtitle).font = SUBTITLE_FONT
        ws.cell(row=4, column=1, value=PROVISIONAL_LABEL).font = PROVISIONAL_FONT
        return 6
    ws.cell(row=3, column=1, value=PROVISIONAL_LABEL).font = PROVISIONAL_FONT
    return 5


def _write_toc_sheet(
    ws: Any,
    scenario_label: str,
    places_df: pd.DataFrame,
    county_names: dict[str, str],
    sheet_lookup: dict[str, str],
    key_years: list[int],
) -> None:
    """Populate Table of Contents with hyperlinks and summary fields."""
    row = _write_header_block(
        ws,
        title="North Dakota Place Projections",
        scenario_label=scenario_label,
        subtitle=f"Generated: {TODAY.strftime('%B %d, %Y')}",
    )

    headers = [
        "Place Name",
        "County",
        "Confidence Tier",
        str(key_years[0]),
        str(key_years[-1]),
        f"% Change ({key_years[0]}-{key_years[-1]})",
        "Sheet",
        "Place FIPS",
    ]
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        cell.font = COL_HEADER_FONT
        cell.fill = COL_HEADER_FILL
        cell.alignment = Alignment(horizontal="center")
    row += 1

    ws.freeze_panes = ws.cell(row=row, column=1)

    tier_rank = {tier: index for index, tier in enumerate(TIER_ORDER)}
    ordered = places_df.copy()
    ordered["tier_rank"] = ordered["confidence_tier"].map(tier_rank)
    ordered = ordered.sort_values(["tier_rank", "name", "place_fips"]).reset_index(drop=True)

    for _, place_row in ordered.iterrows():
        place_fips = str(place_row["place_fips"])
        sheet_name = sheet_lookup[place_fips]
        county_name = county_names.get(str(place_row["county_fips"]), str(place_row["county_fips"]))

        link_cell = ws.cell(row=row, column=1, value=str(place_row["name"]))
        link_cell.font = LINK_FONT
        link_cell.hyperlink = f"#'{sheet_name}'!A1"

        ws.cell(row=row, column=2, value=county_name).font = NORMAL_FONT
        ws.cell(row=row, column=3, value=str(place_row["confidence_tier"])).font = NORMAL_FONT

        ws.cell(row=row, column=4, value=float(place_row["base_population"])).number_format = NUM_FMT
        ws.cell(row=row, column=5, value=float(place_row["final_population"])).number_format = NUM_FMT
        ws.cell(row=row, column=6, value=float(place_row["growth_rate"])).number_format = PCT_FMT

        sheet_cell = ws.cell(row=row, column=7, value=sheet_name)
        sheet_cell.font = LINK_FONT
        sheet_cell.hyperlink = f"#'{sheet_name}'!A1"
        ws.cell(row=row, column=8, value=place_fips).font = NORMAL_FONT

        for col_idx in range(1, len(headers) + 1):
            ws.cell(row=row, column=col_idx).border = THIN_BORDER
        row += 1

    ws.column_dimensions["A"].width = 24
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["C"].width = 15
    ws.column_dimensions["D"].width = 12
    ws.column_dimensions["E"].width = 12
    ws.column_dimensions["F"].width = 18
    ws.column_dimensions["G"].width = 18
    ws.column_dimensions["H"].width = 12


def _write_age_sex_sheet(
    ws: Any,
    *,
    place_name: str,
    place_fips: str,
    county_name: str,
    tier: str,
    scenario_label: str,
    place_projection_df: pd.DataFrame,
    age_groups: list[str],
    key_years: list[int],
) -> None:
    """Write one HIGH or MODERATE place sheet."""
    start_row = _write_header_block(
        ws,
        title=place_name,
        scenario_label=scenario_label,
        subtitle=f"{tier} tier | County: {county_name} | FIPS: {place_fips}",
    )

    ws.cell(row=start_row - 1, column=1, value=f"Tier detail: {tier}").font = SECTION_FONT
    ws.cell(row=start_row - 1, column=6, value="\u2190 Back to Table of Contents").font = LINK_FONT
    ws.cell(row=start_row - 1, column=6).hyperlink = "#'Table of Contents'!A1"

    headers = ["Age Group"]
    for year in key_years:
        headers.extend([f"{year} Male", f"{year} Female"])

    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=start_row, column=col_idx, value=header)
        cell.font = COL_HEADER_FONT
        cell.fill = COL_HEADER_FILL
        cell.alignment = Alignment(horizontal="center")

    row = start_row + 1
    projection_df = place_projection_df.copy()
    if "age_group" not in projection_df.columns or "sex" not in projection_df.columns:
        raise ValueError(f"{place_name} projection missing age_group/sex detail for {tier} sheet")

    grouped = (
        projection_df.groupby(["age_group", "sex", "year"], as_index=False, observed=True)["population"].sum()
    )
    grouped["age_group"] = grouped["age_group"].astype(str)
    grouped["sex"] = grouped["sex"].astype(str)

    values = {
        (str(row_data["age_group"]), str(row_data["sex"]), int(row_data["year"])): float(row_data["population"])
        for _, row_data in grouped.iterrows()
    }

    for age_group in age_groups:
        ws.cell(row=row, column=1, value=age_group).font = NORMAL_FONT
        col = 2
        for year in key_years:
            male = values.get((age_group, "Male", year), 0.0)
            female = values.get((age_group, "Female", year), 0.0)
            ws.cell(row=row, column=col, value=male).number_format = NUM_FMT
            ws.cell(row=row, column=col + 1, value=female).number_format = NUM_FMT
            col += 2
        for col_idx in range(1, len(headers) + 1):
            ws.cell(row=row, column=col_idx).border = THIN_BORDER
        row += 1

    ws.column_dimensions["A"].width = 12
    for col_idx in range(2, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = 11


def _write_lower_combined_sheet(
    ws: Any,
    *,
    lower_df: pd.DataFrame,
    details_by_place: dict[str, pd.DataFrame],
    county_names: dict[str, str],
    scenario_label: str,
    key_years: list[int],
) -> None:
    """Write combined LOWER-tier summary sheet with caveat header."""
    row = _write_header_block(
        ws,
        title="LOWER Tier Places (Combined)",
        scenario_label=scenario_label,
        subtitle="Total population only",
    )

    caveat = (
        "LOWER-tier projections carry wider uncertainty bands and should be used with caution."
    )
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4 + len(key_years))
    caveat_cell = ws.cell(row=row, column=1, value=caveat)
    caveat_cell.font = CAVEAT_FONT
    caveat_cell.fill = PatternFill("solid", fgColor="FFF2CC")
    row += 2

    headers = ["Place Name", "County", "Place FIPS"] + [str(year) for year in key_years]
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        cell.font = COL_HEADER_FONT
        cell.fill = COL_HEADER_FILL
        cell.alignment = Alignment(horizontal="center")
    row += 1

    ordered_lower = lower_df.sort_values(["name", "place_fips"]).reset_index(drop=True)
    for _, lower_row in ordered_lower.iterrows():
        place_fips = str(lower_row["place_fips"])
        county_name = county_names.get(str(lower_row["county_fips"]), str(lower_row["county_fips"]))
        yearly_totals = _totals_by_year(details_by_place[place_fips], key_years)

        ws.cell(row=row, column=1, value=str(lower_row["name"])).font = NORMAL_FONT
        ws.cell(row=row, column=2, value=county_name).font = NORMAL_FONT
        ws.cell(row=row, column=3, value=place_fips).font = NORMAL_FONT
        for year_offset, year in enumerate(key_years, start=4):
            ws.cell(row=row, column=year_offset, value=yearly_totals[year]).number_format = NUM_FMT
        for col_idx in range(1, len(headers) + 1):
            ws.cell(row=row, column=col_idx).border = THIN_BORDER
        row += 1

    ws.column_dimensions["A"].width = 24
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["C"].width = 12
    for col_idx in range(4, 4 + len(key_years)):
        ws.column_dimensions[get_column_letter(col_idx)].width = 11


def _write_methodology_sheet(
    ws: Any,
    *,
    scenario_label: str,
    base_year: int,
    final_year: int,
) -> None:
    """Write methodology and caveat text for place workbook."""
    row = _write_header_block(
        ws,
        title="Methodology",
        scenario_label=scenario_label,
        subtitle="PP-003 place projection methodology summary",
    )

    ws.cell(row=row, column=1, value="Methodology Summary").font = SECTION_FONT
    row += 2

    for line in METHODOLOGY_LINES:
        formatted = line.format(base_year=base_year, final_year=final_year)
        ws.cell(row=row, column=1, value=formatted).font = METHODOLOGY_FONT
        row += 1

    ws.cell(row=row, column=1, value=PLACE_METHODOLOGY_LINE).font = METHODOLOGY_FONT
    row += 2
    ws.cell(row=row, column=1, value=ORGANIZATION_ATTRIBUTION).font = METHODOLOGY_FONT
    row += 1
    ws.cell(row=row, column=1, value=CONDITIONAL_CAVEAT).font = METHODOLOGY_FONT

    ws.column_dimensions["A"].width = 120


HU_COMPARISON_YEARS = [2025, 2030, 2035]
NOTE_FONT = Font(name="Aptos", size=9, italic=True, color="595959")
HU_DIFF_FILL = PatternFill("solid", fgColor="FFF2CC")
PCT_DIFF_FMT = "0.0%"


def _load_hu_projections(scenario: str, config: dict[str, Any]) -> pd.DataFrame | None:
    """
    Load housing-unit projections for a scenario, or return None if unavailable.

    Returns a DataFrame with columns: place_fips, year, population_hu.
    Returns None if the parquet file does not exist or is empty.
    """
    hu_parquet = _projection_root(config) / scenario / "place" / "housing_unit_projections.parquet"
    if not hu_parquet.exists():
        logger.info("HU projections not found for %s (no HU Comparison sheet)", scenario)
        return None

    try:
        hu_df = pd.read_parquet(hu_parquet)
    except Exception:
        logger.warning("Failed to read HU projections: %s", hu_parquet, exc_info=True)
        return None

    if hu_df.empty:
        logger.info("HU projections empty for %s (no HU Comparison sheet)", scenario)
        return None

    required = {"place_fips", "year", "population_hu"}
    if not required.issubset(hu_df.columns):
        logger.warning(
            "HU projections missing columns %s (need %s)",
            required - set(hu_df.columns),
            required,
        )
        return None

    hu_df["place_fips"] = hu_df["place_fips"].astype(str).str.zfill(7)
    hu_df["year"] = pd.to_numeric(hu_df["year"], errors="coerce").astype(int)
    hu_df["population_hu"] = pd.to_numeric(hu_df["population_hu"], errors="coerce")
    return hu_df[["place_fips", "year", "population_hu"]]


def _build_hu_comparison_data(
    places_df: pd.DataFrame,
    details_by_place: dict[str, pd.DataFrame],
    hu_df: pd.DataFrame,
    comparison_years: list[int],
) -> pd.DataFrame | None:
    """
    Build the comparison table between share-trending and HU projections.

    Only includes HIGH and MODERATE tier places that appear in both datasets.
    Returns a DataFrame ready for rendering, or None if no overlap.
    """
    # Filter to HIGH and MODERATE tiers only
    hm_places = places_df[places_df["confidence_tier"].isin(["HIGH", "MODERATE"])].copy()
    if hm_places.empty:
        return None

    # Build share-trending totals for comparison years
    st_rows: list[dict[str, Any]] = []
    for _, place_row in hm_places.iterrows():
        place_fips = str(place_row["place_fips"])
        place_detail = details_by_place.get(place_fips)
        if place_detail is None or place_detail.empty:
            continue
        yearly_totals = _totals_by_year(place_detail, comparison_years)
        for year in comparison_years:
            st_rows.append({
                "place_fips": place_fips,
                "name": str(place_row["name"]),
                "confidence_tier": str(place_row["confidence_tier"]),
                "year": year,
                "pop_share": yearly_totals[year],
            })

    if not st_rows:
        return None

    st_df = pd.DataFrame(st_rows)

    # Filter HU data to comparison years
    hu_subset = hu_df[hu_df["year"].isin(comparison_years)].copy()

    # Merge on place_fips and year — inner join keeps only overlapping places
    merged = st_df.merge(
        hu_subset[["place_fips", "year", "population_hu"]],
        on=["place_fips", "year"],
        how="inner",
    )

    if merged.empty:
        return None

    merged["pop_hu"] = merged["population_hu"]
    merged["abs_diff"] = merged["pop_hu"] - merged["pop_share"]
    merged["pct_diff"] = np.where(
        merged["pop_share"] != 0,
        (merged["pop_hu"] - merged["pop_share"]) / merged["pop_share"],
        np.nan,
    )

    # Sort by tier priority, then name, then year
    tier_rank = {"HIGH": 0, "MODERATE": 1}
    merged["tier_rank"] = merged["confidence_tier"].map(tier_rank)
    merged = merged.sort_values(["tier_rank", "name", "place_fips", "year"]).reset_index(drop=True)
    merged = merged.drop(columns=["tier_rank", "population_hu"])

    return merged


def _write_hu_comparison_sheet(
    ws: Any,
    *,
    comparison_df: pd.DataFrame,
    scenario_label: str,
    comparison_years: list[int],
) -> None:
    """Write the HU Comparison sheet with side-by-side method comparison."""
    row = _write_header_block(
        ws,
        title="Housing-Unit Method Comparison",
        scenario_label=scenario_label,
        subtitle="HIGH and MODERATE tier places",
    )

    # Explanatory note
    note_text = (
        "Note: The housing-unit (HU) method is a complementary cross-check using "
        "housing units \u00d7 persons-per-household (ADR-060). It is NOT the primary "
        "projection method. Differences are expected and reflect methodological "
        "variation, not errors. The share-trending method (ADR-033) remains the "
        "official projection."
    )
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
    note_cell = ws.cell(row=row, column=1, value=note_text)
    note_cell.font = NOTE_FONT
    note_cell.alignment = Alignment(wrap_text=True)
    ws.row_dimensions[row].height = 45
    row += 2

    # Build column headers: one group of columns per comparison year
    headers = ["Place Name", "Tier", "Place FIPS"]
    for year in comparison_years:
        headers.extend([
            f"{year} Share-Trend",
            f"{year} HU Method",
            f"{year} Diff",
            f"{year} % Diff",
        ])

    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        cell.font = COL_HEADER_FONT
        cell.fill = COL_HEADER_FILL
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
    row += 1

    ws.freeze_panes = ws.cell(row=row, column=1)

    # Pivot data to one row per place with year-based columns
    places_in_order = comparison_df.drop_duplicates(
        subset=["place_fips"]
    )[["place_fips", "name", "confidence_tier"]].values.tolist()

    values_lookup: dict[tuple[str, int], dict[str, float]] = {}
    for _, data_row in comparison_df.iterrows():
        key = (str(data_row["place_fips"]), int(data_row["year"]))
        values_lookup[key] = {
            "pop_share": float(data_row["pop_share"]),
            "pop_hu": float(data_row["pop_hu"]),
            "abs_diff": float(data_row["abs_diff"]),
            "pct_diff": float(data_row["pct_diff"]) if pd.notna(data_row["pct_diff"]) else 0.0,
        }

    for place_fips, name, tier in places_in_order:
        ws.cell(row=row, column=1, value=name).font = NORMAL_FONT
        ws.cell(row=row, column=2, value=tier).font = NORMAL_FONT
        ws.cell(row=row, column=3, value=place_fips).font = NORMAL_FONT

        col = 4
        for year in comparison_years:
            vals = values_lookup.get((place_fips, year), {})
            pop_share = vals.get("pop_share", 0.0)
            pop_hu = vals.get("pop_hu", 0.0)
            abs_diff = vals.get("abs_diff", 0.0)
            pct_diff = vals.get("pct_diff", 0.0)

            ws.cell(row=row, column=col, value=pop_share).number_format = NUM_FMT
            ws.cell(row=row, column=col + 1, value=pop_hu).number_format = NUM_FMT

            diff_cell = ws.cell(row=row, column=col + 2, value=abs_diff)
            diff_cell.number_format = NUM_FMT
            if abs(abs_diff) > 0:
                diff_cell.fill = HU_DIFF_FILL

            pct_cell = ws.cell(row=row, column=col + 3, value=pct_diff)
            pct_cell.number_format = PCT_DIFF_FMT
            if abs(pct_diff) > 0:
                pct_cell.fill = HU_DIFF_FILL

            col += 4

        for col_idx in range(1, len(headers) + 1):
            ws.cell(row=row, column=col_idx).border = THIN_BORDER
        row += 1

    # Column widths
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 12
    for col_idx in range(4, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = 14


def build_workbook(
    scenario: str,
    config: dict[str, Any],
    *,
    date_stamp: str = DATE_STAMP,
) -> Path:
    """
    Build place workbook for one scenario and return written file path.

    Args:
        scenario: Scenario key.
        config: Loaded projection configuration dictionary.
        date_stamp: Output file date stamp in YYYYMMDD format.
    """
    scenario_label = SCENARIOS.get(scenario, scenario)
    key_years = _key_years(config)
    base_year = int(key_years[0])
    final_year = int(key_years[-1])

    places_df = _load_places_summary(scenario, config)
    county_names = load_county_names()

    details_by_place: dict[str, pd.DataFrame] = {}
    for place_fips in places_df["place_fips"].astype(str):
        details_by_place[place_fips] = _load_place_projection(scenario, place_fips, config)

    high_df = places_df[places_df["confidence_tier"] == "HIGH"].copy()
    moderate_df = places_df[places_df["confidence_tier"] == "MODERATE"].copy()
    lower_df = places_df[places_df["confidence_tier"] == "LOWER"].copy()

    wb = Workbook()
    toc_ws = wb.active
    toc_ws.title = "Table of Contents"

    used_sheet_names = {"Table of Contents", "LOWER Tier", "HU Comparison", "Methodology"}
    sheet_lookup: dict[str, str] = {}

    for _, row_data in high_df.sort_values(["name", "place_fips"]).iterrows():
        sheet_lookup[str(row_data["place_fips"])] = _unique_sheet_name(
            f"HIGH - {row_data['name']}",
            used_sheet_names,
        )
    for _, row_data in moderate_df.sort_values(["name", "place_fips"]).iterrows():
        sheet_lookup[str(row_data["place_fips"])] = _unique_sheet_name(
            f"MODERATE - {row_data['name']}",
            used_sheet_names,
        )
    for place_fips in lower_df["place_fips"].astype(str):
        sheet_lookup[place_fips] = "LOWER Tier"

    _write_toc_sheet(
        ws=toc_ws,
        scenario_label=scenario_label,
        places_df=places_df,
        county_names=county_names,
        sheet_lookup=sheet_lookup,
        key_years=key_years,
    )

    for _, place_row in high_df.sort_values(["name", "place_fips"]).iterrows():
        place_fips = str(place_row["place_fips"])
        county_name = county_names.get(str(place_row["county_fips"]), str(place_row["county_fips"]))
        ws = wb.create_sheet(sheet_lookup[place_fips])
        _write_age_sex_sheet(
            ws,
            place_name=str(place_row["name"]),
            place_fips=place_fips,
            county_name=county_name,
            tier="HIGH",
            scenario_label=scenario_label,
            place_projection_df=details_by_place[place_fips],
            age_groups=HIGH_AGE_GROUPS,
            key_years=key_years,
        )

    for _, place_row in moderate_df.sort_values(["name", "place_fips"]).iterrows():
        place_fips = str(place_row["place_fips"])
        county_name = county_names.get(str(place_row["county_fips"]), str(place_row["county_fips"]))
        ws = wb.create_sheet(sheet_lookup[place_fips])
        _write_age_sex_sheet(
            ws,
            place_name=str(place_row["name"]),
            place_fips=place_fips,
            county_name=county_name,
            tier="MODERATE",
            scenario_label=scenario_label,
            place_projection_df=details_by_place[place_fips],
            age_groups=MODERATE_AGE_GROUPS,
            key_years=key_years,
        )

    lower_ws = wb.create_sheet("LOWER Tier")
    _write_lower_combined_sheet(
        lower_ws,
        lower_df=lower_df,
        details_by_place=details_by_place,
        county_names=county_names,
        scenario_label=scenario_label,
        key_years=key_years,
    )

    # Optional HU Comparison sheet — only added when HU projection data is available
    hu_df = _load_hu_projections(scenario, config)
    hu_sheet_added = False
    if hu_df is not None:
        comparison_years = [y for y in HU_COMPARISON_YEARS if y in key_years]
        if not comparison_years:
            comparison_years = list(HU_COMPARISON_YEARS)
        comparison_data = _build_hu_comparison_data(
            places_df=places_df,
            details_by_place=details_by_place,
            hu_df=hu_df,
            comparison_years=comparison_years,
        )
        if comparison_data is not None and not comparison_data.empty:
            hu_ws = wb.create_sheet("HU Comparison")
            _write_hu_comparison_sheet(
                hu_ws,
                comparison_df=comparison_data,
                scenario_label=scenario_label,
                comparison_years=comparison_years,
            )
            hu_sheet_added = True
            logger.info(
                "Added HU Comparison sheet: %d places, years %s",
                comparison_data["place_fips"].nunique(),
                comparison_years,
            )

    methodology_ws = wb.create_sheet("Methodology")
    _write_methodology_sheet(
        methodology_ws,
        scenario_label=scenario_label,
        base_year=base_year,
        final_year=final_year,
    )

    export_root = _export_root(config)
    export_root.mkdir(parents=True, exist_ok=True)
    output_path = export_root / f"nd_projections_{scenario}_places_{date_stamp}.xlsx"
    wb.save(output_path)

    logger.info(
        "Built place workbook for %s: %s (sheets=%s, high=%s, moderate=%s, lower=%s, hu_comparison=%s)",
        scenario,
        output_path,
        len(wb.sheetnames),
        len(high_df),
        len(moderate_df),
        len(lower_df),
        hu_sheet_added,
    )
    return output_path


def parse_args() -> argparse.Namespace:
    """Parse CLI arguments."""
    parser = argparse.ArgumentParser(description="Build standalone place workbooks (PP-003 IMP-14)")
    parser.add_argument(
        "--config",
        type=Path,
        default=None,
        help="Path to projection config YAML (default: config/projection_config.yaml)",
    )
    parser.add_argument(
        "--scenarios",
        nargs="+",
        default=None,
        help="Scenario keys to build (default: active scenarios from config)",
    )
    return parser.parse_args()


def main() -> int:
    """CLI entrypoint."""
    logging.basicConfig(level=logging.INFO, format="%(levelname)s: %(message)s")
    args = parse_args()

    try:
        config = load_projection_config(args.config)
        if not isinstance(config, dict):
            raise ValueError("Projection configuration did not load as a dictionary.")

        scenarios = _resolve_scenarios(config, args.scenarios)
        for scenario in scenarios:
            build_workbook(scenario=scenario, config=config)
        return 0
    except Exception as exc:
        logger.error("Failed to build place workbook(s): %s", exc)
        return 1


if __name__ == "__main__":
    sys.exit(main())
