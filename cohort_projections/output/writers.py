"""
Enhanced output writers for projection data.

This module provides enhanced export functionality for population projection results,
supporting multiple formats with rich formatting and customization options.

Functions:
    write_projection_excel: Write projection to formatted Excel workbook
    write_projection_csv: Enhanced CSV export with options
    write_projection_formats: Write to multiple formats at once
    write_projection_shapefile: Export with geographic boundaries (optional)
"""

import pandas as pd
import numpy as np
from pathlib import Path
from typing import Dict, List, Optional, Literal, Any, Union
from datetime import datetime
import json
import warnings

# Excel formatting (openpyxl is optional but recommended)
try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils.dataframe import dataframe_to_rows
    from openpyxl.chart import BarChart, LineChart, Reference
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
    warnings.warn("openpyxl not available - Excel formatting will be limited")

# Geospatial exports (geopandas is optional)
try:
    import geopandas as gpd
    GEOPANDAS_AVAILABLE = True
except ImportError:
    GEOPANDAS_AVAILABLE = False

from ..utils.logger import get_logger_from_config

logger = get_logger_from_config(__name__)


def write_projection_excel(
    projection_df: pd.DataFrame,
    output_path: Union[str, Path],
    summary_df: Optional[pd.DataFrame] = None,
    metadata: Optional[Dict[str, Any]] = None,
    include_charts: bool = True,
    include_formatting: bool = True,
    title: Optional[str] = None
) -> Path:
    """
    Write projection data to a formatted Excel workbook.

    Creates a multi-sheet Excel workbook with:
    - Summary sheet: Key statistics and totals
    - By Age sheet: Population by age group
    - By Sex sheet: Population by sex
    - By Race sheet: Population by race/ethnicity
    - Detail sheet: Full projection data
    - Metadata sheet: Projection parameters and info
    - Charts: Population pyramids and trends (if requested)

    Args:
        projection_df: Full projection DataFrame
                      Required columns: [year, age, sex, race, population]
        output_path: Path to output Excel file (.xlsx)
        summary_df: Optional summary statistics DataFrame
        metadata: Optional metadata dictionary
        include_charts: Whether to include embedded charts
        include_formatting: Whether to apply formatting (colors, borders, etc.)
        title: Optional title for the workbook

    Returns:
        Path to created Excel file

    Raises:
        ValueError: If projection_df is empty or missing required columns
        ImportError: If openpyxl is not available

    Example:
        >>> write_projection_excel(
        ...     projection_df=results,
        ...     output_path='output/projection_2025_2045.xlsx',
        ...     include_charts=True
        ... )
    """
    if not OPENPYXL_AVAILABLE:
        logger.error("openpyxl not available - cannot create formatted Excel file")
        raise ImportError("openpyxl required for Excel export. Install with: pip install openpyxl")

    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    logger.info(f"Creating formatted Excel workbook: {output_path}")

    # Validate input
    required_cols = ['year', 'age', 'sex', 'race', 'population']
    missing_cols = [col for col in required_cols if col not in projection_df.columns]
    if missing_cols:
        raise ValueError(f"projection_df missing required columns: {missing_cols}")

    if projection_df.empty:
        raise ValueError("projection_df is empty")

    # Create workbook
    wb = Workbook()
    wb.remove(wb.active)  # Remove default sheet

    # Define styles
    if include_formatting:
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")

        number_format = '#,##0'
        decimal_format = '#,##0.00'
        percent_format = '0.0%'

        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

    # 1. Summary Sheet
    logger.debug("Creating Summary sheet")
    ws_summary = wb.create_sheet("Summary")

    if summary_df is not None and not summary_df.empty:
        # Write summary data
        for r_idx, row in enumerate(dataframe_to_rows(summary_df, index=False, header=True), 1):
            for c_idx, value in enumerate(row, 1):
                cell = ws_summary.cell(row=r_idx, column=c_idx, value=value)

                if include_formatting:
                    if r_idx == 1:  # Header row
                        cell.font = header_font
                        cell.fill = header_fill
                        cell.alignment = header_alignment
                    elif isinstance(value, (int, float)) and not pd.isna(value):
                        cell.number_format = number_format

        # Auto-width columns
        if include_formatting:
            for column in ws_summary.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                ws_summary.column_dimensions[column_letter].width = min(max_length + 2, 50)

        # Freeze panes
        ws_summary.freeze_panes = 'A2'
    else:
        # Create basic summary from projection data
        years = sorted(projection_df['year'].unique())
        summary_data = []

        for year in years:
            year_data = projection_df[projection_df['year'] == year]
            total_pop = year_data['population'].sum()
            male_pop = year_data[year_data['sex'] == 'Male']['population'].sum()
            female_pop = year_data[year_data['sex'] == 'Female']['population'].sum()

            summary_data.append({
                'Year': int(year),
                'Total Population': int(total_pop),
                'Male': int(male_pop),
                'Female': int(female_pop),
                'Sex Ratio': (male_pop / female_pop * 100) if female_pop > 0 else 0
            })

        summary_df = pd.DataFrame(summary_data)

        # Write to sheet
        for r_idx, row in enumerate(dataframe_to_rows(summary_df, index=False, header=True), 1):
            for c_idx, value in enumerate(row, 1):
                cell = ws_summary.cell(row=r_idx, column=c_idx, value=value)

                if include_formatting:
                    if r_idx == 1:
                        cell.font = header_font
                        cell.fill = header_fill
                        cell.alignment = header_alignment
                    elif c_idx > 1 and isinstance(value, (int, float)):
                        if 'Ratio' in summary_df.columns[c_idx - 1]:
                            cell.number_format = decimal_format
                        else:
                            cell.number_format = number_format

        if include_formatting:
            for col_idx, column in enumerate(ws_summary.columns, 1):
                column_letter = column[0].column_letter
                max_length = len(summary_df.columns[col_idx - 1]) + 2
                ws_summary.column_dimensions[column_letter].width = max_length + 5

        ws_summary.freeze_panes = 'A2'

    # 2. By Age Sheet
    logger.debug("Creating By Age sheet")
    ws_age = wb.create_sheet("By Age")

    # Pivot: years as columns, ages as rows
    age_pivot = projection_df.groupby(['year', 'age'])['population'].sum().reset_index()
    age_pivot = age_pivot.pivot(index='age', columns='year', values='population')
    age_pivot = age_pivot.fillna(0)
    age_pivot.columns = [f'Year {int(col)}' for col in age_pivot.columns]
    age_pivot.index.name = 'Age'
    age_pivot = age_pivot.reset_index()

    # Write to sheet
    for r_idx, row in enumerate(dataframe_to_rows(age_pivot, index=False, header=True), 1):
        for c_idx, value in enumerate(row, 1):
            cell = ws_age.cell(row=r_idx, column=c_idx, value=value)

            if include_formatting:
                if r_idx == 1:
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = header_alignment
                elif c_idx > 1 and isinstance(value, (int, float)):
                    cell.number_format = number_format

    if include_formatting:
        for column in ws_age.columns:
            column_letter = column[0].column_letter
            ws_age.column_dimensions[column_letter].width = 12

    ws_age.freeze_panes = 'B2'

    # 3. By Sex Sheet
    logger.debug("Creating By Sex sheet")
    ws_sex = wb.create_sheet("By Sex")

    sex_data = projection_df.groupby(['year', 'sex'])['population'].sum().reset_index()
    sex_pivot = sex_data.pivot(index='year', columns='sex', values='population')
    sex_pivot['Total'] = sex_pivot.sum(axis=1)
    sex_pivot.index.name = 'Year'
    sex_pivot = sex_pivot.reset_index()

    for r_idx, row in enumerate(dataframe_to_rows(sex_pivot, index=False, header=True), 1):
        for c_idx, value in enumerate(row, 1):
            cell = ws_sex.cell(row=r_idx, column=c_idx, value=value)

            if include_formatting:
                if r_idx == 1:
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = header_alignment
                elif c_idx > 1 and isinstance(value, (int, float)):
                    cell.number_format = number_format

    if include_formatting:
        for column in ws_sex.columns:
            column_letter = column[0].column_letter
            ws_sex.column_dimensions[column_letter].width = 15

    ws_sex.freeze_panes = 'A2'

    # 4. By Race Sheet
    logger.debug("Creating By Race sheet")
    ws_race = wb.create_sheet("By Race")

    race_data = projection_df.groupby(['year', 'race'])['population'].sum().reset_index()
    race_pivot = race_data.pivot(index='year', columns='race', values='population')
    race_pivot['Total'] = race_pivot.sum(axis=1)
    race_pivot.index.name = 'Year'
    race_pivot = race_pivot.reset_index()

    for r_idx, row in enumerate(dataframe_to_rows(race_pivot, index=False, header=True), 1):
        for c_idx, value in enumerate(row, 1):
            cell = ws_race.cell(row=r_idx, column=c_idx, value=value)

            if include_formatting:
                if r_idx == 1:
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = header_alignment
                elif c_idx > 1 and isinstance(value, (int, float)):
                    cell.number_format = number_format

    if include_formatting:
        for column in ws_race.columns:
            column_letter = column[0].column_letter
            ws_race.column_dimensions[column_letter].width = 25

    ws_race.freeze_panes = 'A2'

    # 5. Detail Sheet (full data - limit to 1M rows for Excel)
    logger.debug("Creating Detail sheet")
    ws_detail = wb.create_sheet("Detail")

    # Sort and limit if necessary
    detail_df = projection_df.copy().sort_values(['year', 'age', 'sex', 'race'])
    if len(detail_df) > 1000000:
        logger.warning(f"Detail data has {len(detail_df)} rows, truncating to 1M for Excel")
        detail_df = detail_df.head(1000000)

    for r_idx, row in enumerate(dataframe_to_rows(detail_df, index=False, header=True), 1):
        for c_idx, value in enumerate(row, 1):
            cell = ws_detail.cell(row=r_idx, column=c_idx, value=value)

            if include_formatting:
                if r_idx == 1:
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = header_alignment
                elif 'population' in detail_df.columns and c_idx == list(detail_df.columns).index('population') + 1:
                    if isinstance(value, (int, float)):
                        cell.number_format = decimal_format

    if include_formatting:
        for column in ws_detail.columns:
            column_letter = column[0].column_letter
            ws_detail.column_dimensions[column_letter].width = 15

    ws_detail.freeze_panes = 'A2'

    # 6. Metadata Sheet
    logger.debug("Creating Metadata sheet")
    ws_metadata = wb.create_sheet("Metadata")

    meta_row = 1

    # Title
    if title:
        ws_metadata.cell(row=meta_row, column=1, value=title)
        if include_formatting:
            ws_metadata.cell(row=meta_row, column=1).font = Font(bold=True, size=14)
        meta_row += 2

    # Generation info
    ws_metadata.cell(row=meta_row, column=1, value="Generated")
    ws_metadata.cell(row=meta_row, column=2, value=datetime.now().strftime("%Y-%m-%d %H:%M:%S"))
    meta_row += 1

    ws_metadata.cell(row=meta_row, column=1, value="Software")
    ws_metadata.cell(row=meta_row, column=2, value="North Dakota Cohort Component Projection System")
    meta_row += 2

    # Projection parameters
    if metadata:
        ws_metadata.cell(row=meta_row, column=1, value="Projection Parameters")
        if include_formatting:
            ws_metadata.cell(row=meta_row, column=1).font = Font(bold=True)
        meta_row += 1

        for key, value in metadata.items():
            if isinstance(value, dict):
                # Nested dict - write subsection
                ws_metadata.cell(row=meta_row, column=1, value=f"{key}:")
                if include_formatting:
                    ws_metadata.cell(row=meta_row, column=1).font = Font(bold=True)
                meta_row += 1

                for sub_key, sub_value in value.items():
                    ws_metadata.cell(row=meta_row, column=1, value=f"  {sub_key}")
                    ws_metadata.cell(row=meta_row, column=2, value=str(sub_value))
                    meta_row += 1
            else:
                ws_metadata.cell(row=meta_row, column=1, value=key)
                ws_metadata.cell(row=meta_row, column=2, value=str(value))
                meta_row += 1

    # Column widths
    ws_metadata.column_dimensions['A'].width = 30
    ws_metadata.column_dimensions['B'].width = 50

    # 7. Add Charts (if requested and data available)
    if include_charts and len(projection_df) > 0:
        try:
            logger.debug("Adding charts to Summary sheet")

            # Population trend line chart
            chart = LineChart()
            chart.title = "Total Population Projection"
            chart.style = 10
            chart.y_axis.title = "Population"
            chart.x_axis.title = "Year"

            # Data from summary sheet
            data = Reference(ws_summary, min_col=2, min_row=1, max_row=len(summary_df) + 1, max_col=2)
            cats = Reference(ws_summary, min_col=1, min_row=2, max_row=len(summary_df) + 1)
            chart.add_data(data, titles_from_data=True)
            chart.set_categories(cats)

            ws_summary.add_chart(chart, "F2")

        except Exception as e:
            logger.warning(f"Failed to add charts: {e}")

    # Save workbook
    wb.save(output_path)
    logger.info(f"Excel workbook created successfully: {output_path}")

    return output_path


def write_projection_csv(
    projection_df: pd.DataFrame,
    output_path: Union[str, Path],
    format_type: Literal['wide', 'long'] = 'long',
    age_ranges: Optional[List[tuple]] = None,
    sexes: Optional[List[str]] = None,
    races: Optional[List[str]] = None,
    columns_order: Optional[List[str]] = None,
    compression: Optional[str] = None
) -> Path:
    """
    Write projection data to CSV with enhanced options.

    Supports:
    - Wide format (years as columns) or long format (years as rows)
    - Filtering by age ranges, sexes, races
    - Custom column ordering
    - Optional gzip compression

    Args:
        projection_df: Projection DataFrame
        output_path: Path to output CSV file
        format_type: 'wide' (years as columns) or 'long' (years as rows)
        age_ranges: Optional list of (min_age, max_age) tuples to include
        sexes: Optional list of sexes to include
        races: Optional list of races to include
        columns_order: Optional custom column order
        compression: Optional compression ('gzip', 'bz2', 'zip', 'xz')

    Returns:
        Path to created CSV file

    Example:
        >>> write_projection_csv(
        ...     projection_df=results,
        ...     output_path='output/projection.csv.gz',
        ...     format_type='wide',
        ...     age_ranges=[(0, 17), (18, 64), (65, 90)],
        ...     compression='gzip'
        ... )
    """
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    logger.info(f"Writing CSV: {output_path} (format: {format_type})")

    # Make a copy to avoid modifying original
    df = projection_df.copy()

    # Apply filters
    if age_ranges:
        logger.debug(f"Filtering to age ranges: {age_ranges}")
        age_mask = pd.Series([False] * len(df), index=df.index)
        for min_age, max_age in age_ranges:
            age_mask |= (df['age'] >= min_age) & (df['age'] <= max_age)
        df = df[age_mask]

    if sexes:
        logger.debug(f"Filtering to sexes: {sexes}")
        df = df[df['sex'].isin(sexes)]

    if races:
        logger.debug(f"Filtering to races: {races}")
        df = df[df['race'].isin(races)]

    # Format data
    if format_type == 'wide':
        # Pivot to wide format (years as columns)
        # Group by cohort characteristics
        group_cols = [col for col in ['age', 'sex', 'race'] if col in df.columns]

        df = df.pivot_table(
            index=group_cols,
            columns='year',
            values='population',
            aggfunc='sum'
        ).reset_index()

        # Rename year columns
        df.columns = [f'year_{int(col)}' if isinstance(col, (int, float)) else col
                     for col in df.columns]

    # Apply column ordering
    if columns_order:
        available_cols = [col for col in columns_order if col in df.columns]
        other_cols = [col for col in df.columns if col not in columns_order]
        df = df[available_cols + other_cols]

    # Write to CSV
    df.to_csv(output_path, index=False, compression=compression)

    file_size = output_path.stat().st_size
    logger.info(f"CSV written successfully: {output_path} ({file_size:,} bytes)")

    return output_path


def write_projection_formats(
    projection_df: pd.DataFrame,
    output_dir: Union[str, Path],
    base_filename: str,
    formats: List[Literal['csv', 'excel', 'parquet', 'json']] = ['csv', 'excel', 'parquet'],
    summary_df: Optional[pd.DataFrame] = None,
    metadata: Optional[Dict[str, Any]] = None,
    compression: Optional[str] = 'gzip'
) -> Dict[str, Path]:
    """
    Write projection data to multiple formats at once.

    Exports the same data to CSV, Excel, Parquet, and/or JSON with
    consistent naming and metadata tracking.

    Args:
        projection_df: Projection DataFrame
        output_dir: Output directory
        base_filename: Base filename (without extension)
        formats: List of formats to export ('csv', 'excel', 'parquet', 'json')
        summary_df: Optional summary DataFrame
        metadata: Optional metadata dictionary
        compression: Compression for formats that support it

    Returns:
        Dictionary mapping format -> output path

    Example:
        >>> paths = write_projection_formats(
        ...     projection_df=results,
        ...     output_dir='output/projections',
        ...     base_filename='nd_state_2025_2045',
        ...     formats=['csv', 'excel', 'parquet']
        ... )
        >>> paths['excel']
        PosixPath('output/projections/nd_state_2025_2045.xlsx')
    """
    output_dir = Path(output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)

    logger.info(f"Writing projection to multiple formats: {formats}")

    output_paths = {}

    # CSV
    if 'csv' in formats:
        csv_path = output_dir / f"{base_filename}.csv"
        if compression:
            csv_path = output_dir / f"{base_filename}.csv.gz"

        write_projection_csv(
            projection_df,
            csv_path,
            format_type='long',
            compression=compression if compression else None
        )
        output_paths['csv'] = csv_path

    # Excel
    if 'excel' in formats:
        if OPENPYXL_AVAILABLE:
            excel_path = output_dir / f"{base_filename}.xlsx"
            write_projection_excel(
                projection_df,
                excel_path,
                summary_df=summary_df,
                metadata=metadata,
                include_charts=True
            )
            output_paths['excel'] = excel_path
        else:
            logger.warning("openpyxl not available - skipping Excel export")

    # Parquet
    if 'parquet' in formats:
        parquet_path = output_dir / f"{base_filename}.parquet"
        projection_df.to_parquet(
            parquet_path,
            compression=compression,
            index=False
        )
        output_paths['parquet'] = parquet_path
        logger.info(f"Wrote Parquet: {parquet_path}")

    # JSON
    if 'json' in formats:
        json_path = output_dir / f"{base_filename}.json"

        # Convert to JSON-friendly format
        json_data = {
            'metadata': metadata or {},
            'projection': projection_df.to_dict(orient='records')
        }

        if summary_df is not None:
            json_data['summary'] = summary_df.to_dict(orient='records')

        with open(json_path, 'w') as f:
            json.dump(json_data, f, indent=2, default=str)

        output_paths['json'] = json_path
        logger.info(f"Wrote JSON: {json_path}")

    # Create metadata file
    if metadata:
        metadata_path = output_dir / f"{base_filename}_metadata.json"
        with open(metadata_path, 'w') as f:
            json.dump(metadata, f, indent=2, default=str)
        output_paths['metadata'] = metadata_path
        logger.info(f"Wrote metadata: {metadata_path}")

    logger.info(f"Successfully wrote {len(output_paths)} output files")

    return output_paths


def write_projection_shapefile(
    projection_df: pd.DataFrame,
    geography_level: Literal['state', 'county', 'place'],
    output_path: Union[str, Path],
    year: Optional[int] = None,
    geography_fips: Optional[str] = None,
    format_type: Literal['shapefile', 'geojson'] = 'geojson',
    tiger_vintage: int = 2020
) -> Path:
    """
    Export projection data with geographic boundaries.

    Integrates projection data with Census TIGER boundary files to create
    geospatial outputs suitable for mapping and GIS analysis.

    Args:
        projection_df: Projection DataFrame
        geography_level: Geographic level ('state', 'county', 'place')
        output_path: Path to output file
        year: Optional specific year to export (default: all years)
        geography_fips: Optional specific geography FIPS code
        format_type: Output format ('shapefile' or 'geojson')
        tiger_vintage: Census TIGER vintage year (default: 2020)

    Returns:
        Path to created geospatial file

    Raises:
        ImportError: If geopandas is not available
        ValueError: If no geographic data found

    Note:
        Requires geopandas and Census TIGER boundary files.
        This is an optional feature for advanced geospatial analysis.

    Example:
        >>> write_projection_shapefile(
        ...     projection_df=results,
        ...     geography_level='county',
        ...     output_path='output/counties_2045.geojson',
        ...     year=2045
        ... )
    """
    if not GEOPANDAS_AVAILABLE:
        logger.error("geopandas not available - cannot create geospatial file")
        raise ImportError("geopandas required for shapefile export. Install with: pip install geopandas")

    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    logger.info(f"Creating geospatial output: {output_path}")

    # This is a placeholder implementation
    # Full implementation would require:
    # 1. Loading Census TIGER boundaries for the geography level
    # 2. Joining projection data to boundaries
    # 3. Exporting to shapefile or GeoJSON

    logger.warning("write_projection_shapefile is not fully implemented - requires TIGER boundary data")

    # Basic implementation outline:
    # 1. Load TIGER boundaries (would need to download or have local copies)
    # 2. Filter projection data to specified year
    # 3. Aggregate projection data appropriately
    # 4. Join to geography
    # 5. Export

    raise NotImplementedError(
        "Geospatial export requires Census TIGER boundary files. "
        "Please use write_projection_csv or write_projection_excel for tabular exports."
    )
